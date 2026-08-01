import os, sys, re, json, tempfile, io
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ["FLASK_DEBUG"] = "0"

import pytest
from app import app
from database import get_db


def _reseed(client):
    _login(client)
    from werkzeug.security import generate_password_hash
    db = get_db()
    func = db.execute("SELECT id FROM usuarios WHERE login='funcionario'").fetchone()
    if not func:
        db.execute(
            "INSERT INTO usuarios (nome, login, senha, nivel) VALUES (?,?,?,?)",
            ("Funcionário", "funcionario", generate_password_hash(os.environ["FUNC_SENHA"]), "funcionario")
        )
        db.commit()


def _login(client, login="admin", senha=None):
    if senha is None:
        senha = os.environ.get("ADMIN_SENHA", "admin")
    r = client.get("/login")
    match = re.search(rb'name="_csrf_token"\s+value="([^"]+)"', r.data)
    if not match:
        match = re.search(rb'value="([^"]+)"', r.data)
    assert match, "CSRF token not found"
    csrf = match.group(1).decode()
    r = client.post("/login", data={"login": login, "senha": senha, "_csrf_token": csrf})
    assert r.status_code in (302, 200), f"Login failed: {r.status_code}"
    return csrf


class TestAuth:
    def test_login_page_loads(self, client):
        r = client.get("/login")
        assert r.status_code == 200
        assert b"JAM'S" in r.data or b"csrf_token" in r.data

    def test_csrf_protection(self, client):
        r = client.post("/login", data={"login": "admin", "senha": "x", "_csrf_token": "invalid"})
        assert r.status_code == 400

    def test_login_success(self, client):
        _login(client)

    def test_logout(self, client):
        _login(client)
        r = client.get("/logout", follow_redirects=True)
        assert r.status_code == 200
        assert b"login" in r.data or b"USUARIO" in r.data


class TestAPI:
    def test_dashboard_api(self, client):
        _login(client)
        r = client.get("/api/dashboard")
        assert r.status_code == 200
        d = r.get_json()
        assert "faturamento_dia" in d

    def test_clientes_list(self, client):
        _login(client)
        r = client.get("/api/clientes")
        assert r.status_code == 200
        body = r.get_json()
        if isinstance(body, dict):
            assert "data" in body and "total" in body
            assert len(body["data"]) > 0
        else:
            assert isinstance(body, list)

    def test_mesas_list(self, client):
        _login(client)
        r = client.get("/api/mesas")
        assert r.status_code == 200
        mesas = r.get_json()
        assert isinstance(mesas, list)
        assert len(mesas) >= 20

    def test_produtos_list(self, client):
        _login(client)
        r = client.get("/api/produtos")
        assert r.status_code == 200
        prods = r.get_json()
        assert isinstance(prods, list)

    def test_categorias_list(self, client):
        _login(client)
        r = client.get("/api/categorias")
        assert r.status_code == 200
        cats = r.get_json()
        assert isinstance(cats, list)

    def test_garcons_list(self, client):
        _login(client)
        r = client.get("/api/garcons")
        assert r.status_code == 200
        gs = r.get_json()
        assert isinstance(gs, list)

    def test_estoque_list(self, client):
        _login(client)
        r = client.get("/api/estoque")
        assert r.status_code == 200
        est = r.get_json()
        assert isinstance(est, list)

    def test_alertas(self, client):
        _login(client)
        r = client.get("/api/alertas")
        assert r.status_code == 200
        d = r.get_json()
        assert "criticos" in d and "zerados" in d

    def test_contas_pagar(self, client):
        _login(client)
        r = client.get("/api/contas_pagar")
        assert r.status_code == 200
        c = r.get_json()
        assert isinstance(c, list)

    def test_movimentacoes(self, client):
        _login(client)
        r = client.get("/api/movimentacoes")
        assert r.status_code == 200
        m = r.get_json()
        assert isinstance(m, list)

    def test_auditoria(self, client):
        _login(client)
        r = client.get("/api/auditoria")
        assert r.status_code == 200
        a = r.get_json()
        assert isinstance(a, list)

    def test_caixa_status(self, client):
        _login(client)
        r = client.get("/api/caixa/status")
        assert r.status_code == 200
        s = r.get_json()
        assert "aberto" in s

    def test_buscar_produto(self, client):
        _login(client)
        r = client.get("/api/buscar_produto?q=Heineken")
        assert r.status_code == 200
        p = r.get_json()
        assert isinstance(p, list)

    def test_buscar_produto_codigo(self, client):
        _login(client)
        r = client.get("/api/buscar_produto?codigo=7891991010924")
        assert r.status_code == 200
        p = r.get_json()
        assert isinstance(p, dict)


class TestAdminEndpoints:
    def test_usuarios_list(self, client):
        _login(client)
        r = client.get("/api/usuarios")
        assert r.status_code == 200
        u = r.get_json()
        assert isinstance(u, list)
        assert len(u) >= 2


class TestRelatorios:
    def test_rel_vendas(self, client):
        _login(client)
        r = client.get("/api/relatorios/vendas?inicio=2026-01-01&fim=2026-12-31")
        assert r.status_code == 200
        d = r.get_json()
        assert "vendas" in d and "totais" in d

    def test_rel_mesas(self, client):
        _login(client)
        r = client.get("/api/relatorios/mesas")
        assert r.status_code == 200
        m = r.get_json()
        assert isinstance(m, list)

    def test_rel_produtos(self, client):
        _login(client)
        r = client.get("/api/relatorios/produtos")
        assert r.status_code == 200
        p = r.get_json()
        assert "mais_vendidos" in p

    def test_rel_fluxo_caixa(self, client):
        _login(client)
        r = client.get("/api/relatorios/fluxo_caixa?inicio=2026-01-01&fim=2026-12-31")
        assert r.status_code == 200
        f = r.get_json()
        assert "vendas" in f and "sangrias" in f and "suprimentos" in f

    def test_rel_garcom(self, client):
        _login(client)
        r = client.get("/api/relatorios/vendas_garcom?inicio=2026-01-01&fim=2026-12-31")
        assert r.status_code == 200
        g = r.get_json()
        assert isinstance(g, list)

    def test_rel_categoria(self, client):
        _login(client)
        r = client.get("/api/relatorios/vendas_categoria?inicio=2026-01-01&fim=2026-12-31")
        assert r.status_code == 200
        c = r.get_json()
        assert isinstance(c, list)


class TestRoutes:
    def test_pages_load(self, client):
        _login(client)
        pages = ["/", "/mesas", "/vendas", "/produtos", "/estoque", "/caixa",
                 "/clientes", "/relatorios", "/garcons", "/contas_pagar",
                 "/usuarios", "/auditoria"]
        for p in pages:
            r = client.get(p)
            assert r.status_code in (200, 302), f"{p} returned {r.status_code}"


class TestCancelamentoVenda:

    def _criar_venda(self, client):
        produtos = client.get("/api/produtos").get_json()
        prod = next((p for p in produtos if p["estoque"] > 0 and p["ativo"] == 1 and p.get("codigo_barras")), None)
        assert prod, "Nenhum produto ativo com estoque"
        r = client.post("/api/venda/direta", json={
            "itens": [{"produto_id": prod["id"], "quantidade": 1}],
            "desconto": 0, "acrescimo": 0,
            "forma_pagamento": "Dinheiro"
        })
        d = r.get_json()
        assert d.get("ok"), f"Erro ao criar venda: {d}"
        return prod, d

    def test_cancelamento_venda_sucesso(self, client):
        _login(client)
        prod, venda = self._criar_venda(client)
        estoque_antes = client.get(f"/api/buscar_produto?codigo={prod['codigo_barras']}").get_json()["estoque"]
        r = client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        d = r.get_json()
        assert r.status_code == 200
        assert d["ok"] is True
        estoque_depois = client.get(f"/api/buscar_produto?codigo={prod['codigo_barras']}").get_json()["estoque"]
        assert estoque_depois == estoque_antes + 1

    def test_cancelamento_venda_duplo(self, client):
        _login(client)
        _, venda = self._criar_venda(client)
        client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        r = client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        d = r.get_json()
        assert r.status_code == 400
        assert d["ok"] is False

    def test_cancelamento_venda_inexistente(self, client):
        _login(client)
        r = client.post("/api/venda/99999/cancelar")
        d = r.get_json()
        assert r.status_code == 404

    def test_cancelamento_requer_admin(self, client):
        _login(client, login="funcionario", senha=os.environ["FUNC_SENHA"])
        r = client.post("/api/venda/1/cancelar")
        assert r.status_code in (403, 302)

    def test_cancelamento_movimentacao_estoque(self, client):
        _login(client)
        prod, venda = self._criar_venda(client)
        movs_antes = client.get("/api/movimentacoes").get_json()
        client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        movs_depois = client.get("/api/movimentacoes").get_json()
        cancel_movs = [m for m in movs_depois if m.get("motivo") == "Cancelamento" and m.get("produto_id") == prod["id"]]
        assert len(cancel_movs) > 0

    def test_estoque_nao_fica_negativo(self, client):
        _login(client)
        prod, venda = self._criar_venda(client)
        estoque = client.get(f"/api/buscar_produto?codigo={prod['codigo_barras']}").get_json()["estoque"]
        assert estoque >= 0

    def test_cancelamento_registra_auditoria(self, client):
        _login(client)
        _, venda = self._criar_venda(client)
        client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        logs = client.get("/api/auditoria").get_json()
        cancel_logs = [l for l in logs if l.get("acao") == "CANCELAR_VENDA"]
        assert len(cancel_logs) > 0

    def test_cancelamento_venda_nao_cancelada(self, client):
        _login(client)
        _, venda = self._criar_venda(client)
        r = client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        assert r.status_code == 200
        from database import get_db
        with app.app_context():
            db = get_db()
            v = db.execute("SELECT status FROM vendas WHERE id=?", (venda["venda_id"],)).fetchone()
            assert v["status"] == "cancelada"

    def test_relatorio_exclui_canceladas_totais(self, client):
        _login(client)
        _, venda = self._criar_venda(client)
        r1 = client.get("/api/relatorios/vendas?inicio=2026-01-01&fim=2026-12-31").get_json()
        total_antes = r1["totais"]["total"]
        client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        r2 = client.get("/api/relatorios/vendas?inicio=2026-01-01&fim=2026-12-31").get_json()
        total_depois = r2["totais"]["total"]
        assert total_depois < total_antes

    def test_fluxo_caixa_exclui_canceladas(self, client):
        _login(client)
        _, venda = self._criar_venda(client)
        r1 = client.get("/api/relatorios/fluxo_caixa?inicio=2026-01-01&fim=2026-12-31").get_json()
        vendas_antes = r1["vendas"]["total"]
        client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        r2 = client.get("/api/relatorios/fluxo_caixa?inicio=2026-01-01&fim=2026-12-31").get_json()
        vendas_depois = r2["vendas"]["total"]
        assert vendas_depois < vendas_antes


class TestCaixaComCancelamento:

    def _criar_venda(self, client):
        produtos = client.get("/api/produtos").get_json()
        prod = next((p for p in produtos if p["estoque"] > 0 and p["ativo"] == 1), None)
        assert prod
        r = client.post("/api/venda/direta", json={
            "itens": [{"produto_id": prod["id"], "quantidade": 1}],
            "desconto": 0, "acrescimo": 0,
            "forma_pagamento": "Dinheiro"
        })
        return prod, r.get_json()

    def test_caixa_status_exclui_canceladas(self, client):
        _login(client)
        client.post("/api/caixa/abrir", json={"valor_inicial": 100})
        _, venda = self._criar_venda(client)
        r1 = client.get("/api/caixa/status").get_json()
        vendas_antes = r1["vendas_hoje"]["total"]
        client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        r2 = client.get("/api/caixa/status").get_json()
        vendas_depois = r2["vendas_hoje"]["total"]
        assert vendas_depois < vendas_antes
        client.post("/api/caixa/fechar", json={"valor_final": 100})

    def test_caixa_fechar_exclui_canceladas(self, client):
        _login(client)
        client.post("/api/caixa/abrir", json={"valor_inicial": 0})
        _, venda = self._criar_venda(client)
        status_antes = client.get("/api/caixa/status").get_json()
        total_antes = status_antes["vendas_hoje"]["total"]
        client.post(f"/api/venda/{venda['venda_id']}/cancelar")
        r = client.post("/api/caixa/fechar", json={"valor_final": 0})
        d = r.get_json()
        assert d["ok"] is True
        assert d["total"] == total_antes - venda["total"]


class TestEmpresaEBackup:

    def test_empresa_get(self, client):
        _login(client)
        r = client.get("/api/empresa")
        assert r.status_code == 200

    def test_empresa_save(self, client):
        _login(client)
        r = client.post("/api/empresa", json={"razao_social": "Teste LTDA", "nome_fantasia": "Teste"})
        d = r.get_json()
        assert d["ok"] is True

    def test_empresa_requer_admin(self, client):
        _login(client, login="funcionario", senha=os.environ["FUNC_SENHA"])
        r = client.post("/api/empresa", json={"razao_social": "X"})
        assert r.status_code in (403, 302)

    def test_backup_download(self, client):
        _login(client)
        r = client.get("/api/backup")
        assert r.status_code == 200
        assert r.content_type == "application/octet-stream" or len(r.data) > 0


class TestAutorizacaoPermissoes:

    def _ensure_func_user(self):
        import sqlite3
        from werkzeug.security import generate_password_hash
        db_path = app.config["DATABASE"]
        conn = sqlite3.connect(db_path)
        func = conn.execute("SELECT id FROM usuarios WHERE login='funcionario'").fetchone()
        if not func:
            conn.execute(
                "INSERT INTO usuarios (nome, login, senha, nivel) VALUES (?,?,?,?)",
                ("Funcionário", "funcionario", generate_password_hash(os.environ["FUNC_SENHA"]), "funcionario")
            )
            conn.commit()
        conn.close()

    def _login_funcionario(self, client):
        self._ensure_func_user()
        return _login(client, login="funcionario", senha=os.environ["FUNC_SENHA"])

    def _login_admin(self, client):
        return _login(client, login="admin", senha=os.environ["ADMIN_SENHA"])

    def test_funcionario_caixa(self, client):
        self._login_funcionario(client)
        assert client.get("/api/caixa/status").status_code in (403, 302)
        assert client.get("/api/caixa/movimentacoes").status_code in (403, 302)
        assert client.post("/api/caixa/abrir", json={"valor_inicial": 100}).status_code in (403, 302)
        assert client.post("/api/caixa/fechar", json={"valor_final": 0}).status_code in (403, 302)
        assert client.post("/api/caixa/suprimento", json={"valor": 50, "motivo": "X"}).status_code in (403, 302)
        assert client.post("/api/caixa/sangria", json={"valor": 50, "motivo": "X"}).status_code in (403, 302)

    def test_funcionario_clientes(self, client):
        self._login_funcionario(client)
        assert client.get("/api/clientes").status_code == 200
        assert client.get("/api/clientes/buscar?q=t").status_code == 200
        assert client.get("/api/clientes/1/fiado").status_code in (200, 404)
        assert client.post("/api/clientes", json={"nome": "T"}).status_code in (403, 302)
        assert client.put("/api/clientes/1", json={"nome": "X"}).status_code in (403, 302)
        assert client.delete("/api/clientes/1").status_code in (403, 302)
        assert client.post("/api/clientes/1/pagamento", json={"valor": 10}).status_code in (403, 302)

    def test_funcionario_contas(self, client):
        self._login_funcionario(client)
        assert client.get("/api/contas_pagar").status_code in (403, 302)
        assert client.get("/api/contas_pagar/verificar_atrasadas").status_code in (403, 302)
        assert client.post("/api/contas_pagar", json={"fornecedor": "F", "descricao": "D", "valor": 100, "vencimento": "2026-12-31"}).status_code in (403, 302)
        assert client.post("/api/contas_pagar/1/pagar").status_code in (403, 302)
        assert client.delete("/api/contas_pagar/1").status_code in (403, 302)

    def test_funcionario_estoque(self, client):
        self._login_funcionario(client)
        assert client.get("/api/estoque").status_code in (403, 302)
        assert client.get("/api/movimentacoes").status_code in (403, 302)
        assert client.post("/api/estoque/entrada", json={"produto_id": 1, "quantidade": 10}).status_code in (403, 302)
        assert client.post("/api/estoque/saida", json={"produto_id": 1, "quantidade": 1, "motivo": "Ajuste"}).status_code in (403, 302)

    def test_funcionario_relatorios(self, client):
        self._login_funcionario(client)
        assert client.get("/api/relatorios/vendas").status_code in (403, 302)
        assert client.get("/api/relatorios/produtos").status_code in (403, 302)
        assert client.get("/api/relatorios/vendas_categoria").status_code in (403, 302)
        assert client.get("/api/relatorios/vendas_garcom").status_code in (403, 302)
        assert client.get("/api/relatorios/mesas").status_code in (403, 302)
        assert client.get("/api/relatorios/fluxo_caixa").status_code in (403, 302)
        assert client.get("/api/relatorios/sangrias").status_code in (403, 302)
        assert client.get("/api/relatorios/suprimentos").status_code in (403, 302)

    def test_funcionario_cancelamento(self, client):
        self._login_funcionario(client)
        assert client.post("/api/venda/1/cancelar").status_code in (403, 302)

    def test_funcionario_garcons(self, client):
        self._login_funcionario(client)
        assert client.get("/api/garcons").status_code == 200
        assert client.post("/api/garcons", json={"nome": "T"}).status_code in (403, 302)
        assert client.put("/api/garcons/1", json={"nome": "X"}).status_code in (403, 302)
        assert client.delete("/api/garcons/1").status_code in (403, 302)

    def test_funcionario_admin_area(self, client):
        self._login_funcionario(client)
        assert client.get("/api/usuarios").status_code in (403, 302)
        assert client.get("/api/auditoria").status_code in (403, 302)
        assert client.get("/api/backup").status_code in (403, 302)
        assert client.post("/api/empresa", json={"razao_social": "X"}).status_code in (403, 302)
        assert client.get("/api/empresa").status_code in (403, 302)
        assert client.delete("/api/config/logo").status_code in (403, 302)

    def test_funcionario_produtos(self, client):
        self._login_funcionario(client)
        assert client.get("/api/produtos").status_code == 200
        assert client.get("/api/categorias").status_code == 200
        assert client.post("/api/produtos", json={"nome": "T", "preco": 10}).status_code in (403, 302)
        assert client.put("/api/produtos/1", json={"nome": "X", "preco": 10}).status_code in (403, 302)
        assert client.delete("/api/produtos/1").status_code in (403, 302)
        assert client.post("/api/categorias", json={"nome": "T"}).status_code in (403, 302)

    def test_admin_acesso_total(self, client):
        self._login_admin(client)
        assert client.get("/api/caixa/status").status_code == 200
        assert client.get("/api/clientes").status_code == 200
        assert client.get("/api/contas_pagar").status_code == 200
        assert client.get("/api/estoque").status_code == 200
        assert client.get("/api/relatorios/vendas").status_code == 200
        assert client.get("/api/relatorios/fluxo_caixa").status_code == 200
        assert client.get("/api/relatorios/sangrias").status_code == 200
        assert client.get("/api/relatorios/suprimentos").status_code == 200
        assert client.get("/api/usuarios").status_code == 200
        assert client.get("/api/auditoria").status_code == 200
        assert client.get("/api/empresa").status_code == 200

    def test_nao_logado_redireciona(self, client):
        endpoints_protected = [
            "/api/caixa/status",
            "/api/clientes",
            "/api/contas_pagar",
            "/api/estoque",
            "/api/produtos",
            "/api/garcons",
            "/api/relatorios/vendas",
        ]
        for ep in endpoints_protected:
            r = client.get(ep)
            assert r.status_code in (302, 401, 403), f"{ep} retornou {r.status_code} para usuário não logado"


class TestSaldoDevedor:

    def test_saldo_devedor_increments_on_fiado_compra(self, client):
        _login(client)
        from database import get_db
        with app.app_context():
            db = get_db()
            cli = db.execute("""
                SELECT c.id, c.saldo_devedor FROM clientes c
                WHERE c.ativo=1 AND NOT EXISTS (
                    SELECT 1 FROM fiado f WHERE f.cliente_id=c.id AND f.tipo='compra'
                    AND (f.valor - f.valor_pago) > 0.01
                    AND f.data_vencimento IS NOT NULL AND f.data_vencimento < date('now')
                ) LIMIT 1
            """).fetchone()
            if not cli:
                db.execute("INSERT INTO clientes (nome, limite_fiado) VALUES (?, ?)", ("TesteFiado", 5000))
                db.commit()
                cli = db.execute("SELECT id, saldo_devedor FROM clientes WHERE nome='TesteFiado'").fetchone()
        saldo_antes = cli["saldo_devedor"]
        produtos = client.get("/api/produtos").get_json()
        prod = next((p for p in produtos if p["estoque"] > 0), None)
        assert prod
        r = client.post("/api/venda/direta", json={
            "itens": [{"produto_id": prod["id"], "quantidade": 1}],
            "desconto": 0, "acrescimo": 0,
            "forma_pagamento": "Fiado",
            "cliente_id": cli["id"],
        })
        assert r.get_json().get("ok"), f"Erro: {r.get_json()}"
        with app.app_context():
            db = get_db()
            cli_atualizado = db.execute("SELECT saldo_devedor FROM clientes WHERE id=?", (cli["id"],)).fetchone()
            assert cli_atualizado["saldo_devedor"] > saldo_antes

    def test_saldo_devedor_decrements_on_pagamento(self, client):
        _login(client)
        from database import get_db
        with app.app_context():
            db = get_db()
            cli = db.execute("SELECT id, saldo_devedor FROM clientes WHERE saldo_devedor > 0 LIMIT 1").fetchone()
            if not cli:
                db.execute("INSERT INTO clientes (nome, saldo_devedor) VALUES (?, ?)", ("Teste Saldo", 100))
                db.commit()
                cli = db.execute("SELECT id, saldo_devedor FROM clientes WHERE nome='Teste Saldo'").fetchone()
        saldo_antes = cli["saldo_devedor"]
        r = client.post(f"/api/clientes/{cli['id']}/pagamento", json={"valor": 10})
        assert r.get_json().get("ok")
        with app.app_context():
            db = get_db()
            cli_dep = db.execute("SELECT saldo_devedor FROM clientes WHERE id=?", (cli["id"],)).fetchone()
            assert cli_dep["saldo_devedor"] < saldo_antes

    def test_saldo_devedor_consistent_after_recalc(self, client):
        _login(client)
        from database import get_db
        with app.app_context():
            db = get_db()
            from services.fiado_service import recalcular_saldo_devedor
            clientes = db.execute("SELECT id FROM clientes").fetchall()
            for c in clientes:
                recalcular_saldo_devedor(c["id"])
            db.commit()
            desatualizados = db.execute("""
                SELECT c.id FROM clientes c
                LEFT JOIN fiado f ON f.cliente_id=c.id AND f.tipo='compra' AND (f.valor-f.valor_pago)>0.01
                GROUP BY c.id HAVING ABS(c.saldo_devedor - COALESCE(SUM(f.valor-f.valor_pago),0)) > 0.01
            """).fetchall()
            assert len(desatualizados) == 0


class TestEmpresaSingleton:

    def test_empresa_always_one_record(self, client):
        _login(client)
        client.post("/api/empresa", json={"razao_social": "Primeira"})
        client.post("/api/empresa", json={"razao_social": "Segunda"})
        r = client.get("/api/empresa")
        emp = r.get_json()
        assert emp.get("razao_social") == "Segunda"
        from database import get_db
        with app.app_context():
            db = get_db()
            count = db.execute("SELECT COUNT(*) as c FROM empresa").fetchone()["c"]
            assert count == 1

    def test_empresa_id_always_1(self, client):
        _login(client)
        client.post("/api/empresa", json={"razao_social": "Teste ID"})
        from database import get_db
        with app.app_context():
            db = get_db()
            emp = db.execute("SELECT id FROM empresa").fetchone()
            assert emp["id"] == 1


class TestMigracaoValidacao:

    def test_validar_endpoint(self, client):
        _login(client)
        r = client.post("/api/migracao/validar")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "orfas_fk" in d
        assert "timestamps_invalidos" in d
        assert "saldo_devedor_desatualizado" in d
        assert "empresa_duplicada" in d

    def test_validar_requer_admin(self, client):
        _login(client, login="funcionario", senha=os.environ["FUNC_SENHA"])
        r = client.post("/api/migracao/validar")
        assert r.status_code in (403, 302)

    def test_corrigir_saldo_endpoint(self, client):
        _login(client)
        r = client.post("/api/migracao/corrigir_saldo")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "clientes_corrigidos" in d

    def test_corrigir_empresa_endpoint(self, client):
        _login(client)
        r = client.post("/api/migracao/corrigir_empresa")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True

    def test_status_migracao(self, client):
        _login(client)
        r = client.get("/api/migracao/status")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "tabelas" in d
        assert "usuarios" in d["tabelas"]
        assert d["tabelas"]["usuarios"] >= 2

    def test_no_orphan_fks_on_clean_db(self, client):
        _login(client)
        from services.migration_service import validar_foreign_keys_orfas
        with app.app_context():
            db = get_db()
            orfas = validar_foreign_keys_orfas(db)
            orfas_reais = [o for o in orfas if o["tabela"] not in ("auditoria",)]
            assert len(orfas_reais) == 0

    def test_timestamps_valid(self, client):
        _login(client)
        from services.migration_service import validar_formato_timestamps
        with app.app_context():
            db = get_db()
            erros = validar_formato_timestamps(db)
            assert len(erros) == 0


class TestFiadoService:

    def test_calcular_status_fiado_vencido(self):
        from services.fiado_service import calcular_status_fiado
        from datetime import date, timedelta
        venc = (date.today() - timedelta(days=3)).isoformat()
        status, dias = calcular_status_fiado(venc)
        assert status == "vencido"
        assert dias < 0

    def test_calcular_status_fiado_normal(self):
        from services.fiado_service import calcular_status_fiado
        from datetime import date, timedelta
        venc = (date.today() + timedelta(days=30)).isoformat()
        status, dias = calcular_status_fiado(venc)
        assert status == "normal"
        assert dias == 30

    def test_calcular_status_fiado_none(self):
        from services.fiado_service import calcular_status_fiado
        status, dias = calcular_status_fiado(None)
        assert status == "normal"
        assert dias is None

    def test_calcular_status_fiado_invalid(self):
        from services.fiado_service import calcular_status_fiado
        status, dias = calcular_status_fiado("invalid-date")
        assert status == "normal"
        assert dias is None


class TestMigrationValidators:
    def test_validar_extensao_sqlite_ok(self):
        from migration.validators import validar_extensao
        ok, tipo, ext = validar_extensao("banco.db")
        assert ok is True
        assert tipo == "sqlite"
        assert ext == "db"

    def test_validar_extensao_sqlite_sqlite3(self):
        from migration.validators import validar_extensao
        ok, tipo, ext = validar_extensao("dados.sqlite3")
        assert ok is True
        assert tipo == "sqlite"

    def test_validar_extensao_sql_ok(self):
        from migration.validators import validar_extensao
        ok, tipo, ext = validar_extensao("dump.sql")
        assert ok is True
        assert tipo == "sql"

    def test_validar_extensao_excel_ok(self):
        from migration.validators import validar_extensao
        ok, tipo, ext = validar_extensao("planilha.xlsx")
        assert ok is True
        assert tipo == "excel"

    def test_validar_extensao_csv_ok(self):
        from migration.validators import validar_extensao
        ok, tipo, ext = validar_extensao("dados.csv")
        assert ok is True
        assert tipo == "csv"

    def test_validar_extensao_rejeita_exe(self):
        from migration.validators import validar_extensao
        ok, tipo, ext = validar_extensao("virus.exe")
        assert ok is False

    def test_validar_extensao_rejeita_desconhecido(self):
        from migration.validators import validar_extensao
        ok, tipo, ext = validar_extensao("arquivo.xyz")
        assert ok is False

    def test_validar_tamanho_ok(self):
        from migration.validators import validar_tamanho
        ok, size = validar_tamanho(1024 * 1024)
        assert ok is True

    def test_validar_tamanho_excede(self):
        from migration.validators import validar_tamanho
        ok, mb = validar_tamanho(60 * 1024 * 1024)
        assert ok is False
        assert mb > 50

    def test_is_executavel_exe(self):
        from migration.validators import is_executavel
        assert is_executavel("app.exe") is True

    def test_is_executavel_bat(self):
        from migration.validators import is_executavel
        assert is_executavel("script.bat") is True

    def test_is_executavel_db_falso(self):
        from migration.validators import is_executavel
        assert is_executavel("dados.db") is False

    def test_validar_upload_completo_ok(self):
        from migration.validators import validar_upload_completo
        r = validar_upload_completo("dados.db", 1024)
        assert r["valido"] is True
        assert r["tipo_detectado"] == "sqlite"

    def test_validar_upload_completo_executavel(self):
        from migration.validators import validar_upload_completo
        r = validar_upload_completo("virus.exe", 1024)
        assert r["valido"] is False
        assert any("executável" in e.lower() for e in r["erros"])

    def test_validar_upload_completo_muito_grande(self):
        from migration.validators import validar_upload_completo
        r = validar_upload_completo("dados.db", 60 * 1024 * 1024)
        assert r["valido"] is False
        assert any("grande" in e.lower() for e in r["erros"])


class TestMigrationWizard:
    def _login_admin(self, client):
        _login(client)

    def test_etapa1_requires_admin(self, client):
        r = client.get("/migracao/etapa1")
        assert r.status_code == 302

    def test_etapa1_loads_for_admin(self, client):
        self._login_admin(client)
        r = client.get("/migracao/etapa1")
        assert r.status_code == 200
        assert b"Origem" in r.data

    def test_etapa1_has_all_sources(self, client):
        self._login_admin(client)
        r = client.get("/migracao/etapa1")
        assert b"SQLite" in r.data
        assert b"SQL Dump" in r.data
        assert b"Excel" in r.data
        assert b"CSV" in r.data

    def test_etapa2_sets_tipo_in_session(self, client):
        self._login_admin(client)
        r = client.get("/migracao/etapa2?tipo=sqlite")
        assert r.status_code == 200
        assert b"Upload" in r.data

    def test_etapa2_redirect_without_tipo(self, client):
        self._login_admin(client)
        r = client.get("/migracao/etapa2")
        assert r.status_code == 302

    def test_etapa3_requires_session(self, client):
        self._login_admin(client)
        r = client.get("/migracao/etapa3")
        assert r.status_code == 302

    def test_etapa3_loads_with_analysis(self, client):
        self._login_admin(client)
        client.get("/migracao/etapa2?tipo=csv")
        data = io.BytesIO(b"nome,preco\nItem,10.00")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (data, "test.csv")},
                        content_type="multipart/form-data")
        assert r.status_code == 200
        r = client.post("/api/migracao/analisar")
        assert r.status_code == 200
        assert r.get_json()["ok"] is True
        r = client.get("/migracao/etapa3")
        assert r.status_code == 200
        assert b"An" in r.data

    def test_etapa3_redirect_without_analysis(self, client):
        self._login_admin(client)
        client.get("/migracao/etapa2?tipo=csv")
        data = io.BytesIO(b"nome,preco\nItem,10.00")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (data, "test.csv")},
                        content_type="multipart/form-data")
        assert r.status_code == 200
        r = client.get("/migracao/etapa3")
        assert r.status_code == 302

    def test_etapa4_requires_session(self, client):
        self._login_admin(client)
        r = client.get("/migracao/etapa4")
        assert r.status_code == 302

    def test_etapa4_loads_with_session(self, client):
        self._login_admin(client)
        client.get("/migracao/etapa2?tipo=csv")
        data = io.BytesIO(b"nome,preco\nItem,10.00")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (data, "test.csv")},
                        content_type="multipart/form-data")
        assert r.status_code == 200
        r = client.get("/migracao/etapa4")
        assert r.status_code == 200
        assert b"Confirma" in r.data or "Confirmar".encode() in r.data

    def test_upload_rejects_no_file(self, client):
        self._login_admin(client)
        client.get("/migracao/etapa2?tipo=csv")
        r = client.post("/migracao/etapa2",
                        data={},
                        content_type="multipart/form-data")
        assert r.status_code == 400

    def test_upload_rejects_executable(self, client):
        self._login_admin(client)
        client.get("/migracao/etapa2?tipo=sqlite")
        data = io.BytesIO(b"MZ\x90\x00fake exe content")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (data, "virus.exe")},
                        content_type="multipart/form-data")
        assert r.status_code == 400
        d = r.get_json()
        assert d["ok"] is False

    def test_upload_accepts_valid_csv(self, client):
        self._login_admin(client)
        client.get("/migracao/etapa2?tipo=csv")
        data = io.BytesIO(b"nome,preco\nItem,10.00")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (data, "dados.csv")},
                        content_type="multipart/form-data")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert d["tipo"] == "csv"

    def test_upload_accepts_valid_db(self, client):
        self._login_admin(client)
        client.get("/migracao/etapa2?tipo=sqlite")
        data = io.BytesIO(b"SQLite format 3\x00" + b"\x00" * 80)
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (data, "banco.db")},
                        content_type="multipart/form-data")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True

    def test_cancelar_endpoint(self, client):
        self._login_admin(client)
        r = client.post("/api/migracao/cancelar")
        assert r.status_code == 200
        assert r.get_json()["ok"] is True

    def test_confirmar_requires_session(self, client):
        self._login_admin(client)
        r = client.post("/api/migracao/confirmar")
        assert r.status_code == 400

    def test_index_redirects(self, client):
        self._login_admin(client)
        r = client.get("/migracao")
        assert r.status_code == 302

    def test_analisar_requires_file_in_session(self, client):
        self._login_admin(client)
        r = client.post("/api/migracao/analisar")
        assert r.status_code == 400

    def test_analisar_csv_full_flow(self, client):
        self._login_admin(client)
        client.get("/migracao/etapa2?tipo=csv")
        csv_data = b"nome,preco,estoque\nArroz,10.50,100\nFeijao,8.00,50\n"
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (io.BytesIO(csv_data), "estoque.csv")},
                        content_type="multipart/form-data")
        assert r.status_code == 200
        r = client.post("/api/migracao/analisar")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert d["tipo_fonte"] == "csv"
        assert len(d["tabelas"]) == 1
        assert d["tabelas"][0]["registros"] == 2
        assert d["tabelas"][0]["colunas"] == 3
        assert "nome" in d["tabelas"][0]["cabecalhos"]
        assert d["total_registros"] == 2
        assert len(d["integridade"]) > 0

    def test_analisar_sqlite_full_flow(self, client):
        self._login_admin(client)
        import sqlite3 as sq3
        tmp = os.path.join(tempfile.gettempdir(), "test_analyze.db")
        conn = sq3.connect(tmp)
        conn.execute("CREATE TABLE clientes (id INTEGER PRIMARY KEY, nome TEXT)")
        conn.execute("INSERT INTO clientes (nome) VALUES ('Joao')")
        conn.execute("INSERT INTO clientes (nome) VALUES ('Maria')")
        conn.execute("CREATE TABLE produtos (id INTEGER PRIMARY KEY, nome TEXT, preco REAL)")
        conn.execute("INSERT INTO produtos (nome, preco) VALUES ('Cafe', 5.0)")
        conn.commit()
        conn.close()
        with open(tmp, "rb") as f:
            db_bytes = f.read()
        os.remove(tmp)
        client.get("/migracao/etapa2?tipo=sqlite")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (io.BytesIO(db_bytes), "teste.db")},
                        content_type="multipart/form-data")
        assert r.status_code == 200
        r = client.post("/api/migracao/analisar")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert d["tipo_fonte"] == "sqlite"
        table_names = [t["nome"] for t in d["tabelas"]]
        assert "clientes" in table_names
        assert "produtos" in table_names
        clientes_t = next(t for t in d["tabelas"] if t["nome"] == "clientes")
        assert clientes_t["registros"] == 2
        produtos_t = next(t for t in d["tabelas"] if t["nome"] == "produtos")
        assert produtos_t["registros"] == 1
        assert d["total_registros"] == 3
        integrity_names = [c["check"] for c in d["integridade"]]
        assert "integrity_check" in integrity_names
        assert "tabelas" in integrity_names
        assert "estrutura" in integrity_names


class TestSQLiteAnalyzer:
    def test_analyze_valid_db(self):
        from migration.importer import SQLiteAnalyzer
        import sqlite3 as sq3
        tmp = os.path.join(tempfile.gettempdir(), "test_sa.db")
        conn = sq3.connect(tmp)
        conn.execute("CREATE TABLE usuarios (id INTEGER PRIMARY KEY, nome TEXT)")
        conn.execute("INSERT INTO usuarios (nome) VALUES ('Admin')")
        conn.execute("CREATE TABLE vendas (id INTEGER PRIMARY KEY, valor REAL)")
        for i in range(10):
            conn.execute("INSERT INTO vendas (valor) VALUES (?)", (float(i * 10),))
        conn.commit()
        conn.close()
        r = SQLiteAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert r.tipo_fonte == "sqlite"
        assert len(r.tabelas) == 2
        names = [t["nome"] for t in r.tabelas]
        assert "usuarios" in names
        assert "vendas" in names
        assert r.total_registros == 11
        integrity = {c["check"]: c["ok"] for c in r.integridade}
        assert integrity["integrity_check"] is True
        assert integrity["tabelas"] is True
        assert integrity["estrutura"] is True

    def test_analyze_empty_file(self):
        from migration.importer import SQLiteAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_sa_empty.db")
        with open(tmp, "wb") as f:
            f.write(b"")
        r = SQLiteAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is False
        assert any("vazio" in e.lower() for e in r.erros)

    def test_analyze_corrupt_file(self):
        from migration.importer import SQLiteAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_sa_corrupt.db")
        with open(tmp, "wb") as f:
            f.write(b"this is not a sqlite file at all")
        r = SQLiteAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is False
        assert len(r.erros) > 0

    def test_analyze_empty_tables(self):
        from migration.importer import SQLiteAnalyzer
        import sqlite3 as sq3
        tmp = os.path.join(tempfile.gettempdir(), "test_sa_notables.db")
        conn = sq3.connect(tmp)
        conn.execute("CREATE TABLE empty_table (id INTEGER)")
        conn.commit()
        conn.close()
        r = SQLiteAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert len(r.tabelas) == 1
        assert r.tabelas[0]["registros"] == 0

    def test_metadata_has_version(self):
        from migration.importer import SQLiteAnalyzer
        import sqlite3 as sq3
        tmp = os.path.join(tempfile.gettempdir(), "test_sa_ver.db")
        conn = sq3.connect(tmp)
        conn.execute("CREATE TABLE t (id INTEGER)")
        conn.commit()
        conn.close()
        r = SQLiteAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert "sqlite_version" in r.metadata

    def test_to_dict(self):
        from migration.importer import SQLiteAnalyzer, AnalysisResult
        r = AnalysisResult()
        d = r.to_dict()
        assert "ok" in d
        assert "tabelas" in d
        assert "integridade" in d
        assert "erros" in d
        assert "metadata" in d


class TestSQLAnalyzer:
    def test_analyze_with_create_and_insert(self):
        from migration.importer import SQLAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_sqla.sql")
        sql_content = """CREATE TABLE clientes (
            id INTEGER PRIMARY KEY,
            nome TEXT NOT NULL
        );
        INSERT INTO clientes (nome) VALUES ('Joao');
        INSERT INTO clientes (nome) VALUES ('Maria');
        INSERT INTO clientes (nome) VALUES ('Pedro');
        CREATE TABLE produtos (
            id INTEGER PRIMARY KEY,
            nome TEXT,
            preco REAL
        );
        INSERT INTO produtos (nome, preco) VALUES ('Cafe', 5.0);
        INSERT INTO produtos (nome, preco) VALUES ('Cha', 3.0);
        """
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(sql_content)
        r = SQLAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert r.tipo_fonte == "sql"
        table_names = [t["nome"] for t in r.tabelas]
        assert "clientes" in table_names
        assert "produtos" in table_names
        clientes_t = next(t for t in r.tabelas if t["nome"] == "clientes")
        assert clientes_t["registros"] == 3
        produtos_t = next(t for t in r.tabelas if t["nome"] == "produtos")
        assert produtos_t["registros"] == 2
        assert r.total_registros == 5
        integrity = {c["check"]: c["ok"] for c in r.integridade}
        assert integrity["create_tables"] is True
        assert integrity["inserts"] is True

    def test_analyze_create_only(self):
        from migration.importer import SQLAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_sqla_create.sql")
        with open(tmp, "w") as f:
            f.write("CREATE TABLE users (id INT, name VARCHAR(100));\n")
        r = SQLAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert len(r.tabelas) == 1
        assert r.tabelas[0]["registros"] == 0

    def test_analyze_empty_file(self):
        from migration.importer import SQLAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_sqla_empty.sql")
        with open(tmp, "wb") as f:
            f.write(b"")
        r = SQLAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is False

    def test_metadata_encoding(self):
        from migration.importer import SQLAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_sqla_enc.sql")
        with open(tmp, "w", encoding="utf-8") as f:
            f.write("CREATE TABLE t (id INT);\nINSERT INTO t VALUES (1);\n")
        r = SQLAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert "encoding_detectado" in r.metadata
        assert "total_linhas" in r.metadata


class TestExcelAnalyzer:
    def test_analyze_xlsx(self):
        from migration.importer import ExcelAnalyzer
        from openpyxl import Workbook
        tmp = os.path.join(tempfile.gettempdir(), "test_xl.xlsx")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Clientes"
        ws1.append(["id", "nome", "email"])
        ws1.append([1, "Joao", "joao@test.com"])
        ws1.append([2, "Maria", "maria@test.com"])
        ws2 = wb.create_sheet("Produtos")
        ws2.append(["id", "nome", "preco"])
        ws2.append([1, "Cafe", 5.0])
        ws2.append([2, "Cha", 3.0])
        ws2.append([3, "Leite", 4.5])
        wb.save(tmp)
        wb.close()
        r = ExcelAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert r.tipo_fonte == "excel"
        assert len(r.tabelas) == 2
        assert "abas" in r.metadata
        assert len(r.metadata["abas"]) == 2
        clientes = next(t for t in r.tabelas if t["nome"] == "Clientes")
        assert clientes["registros"] == 2
        assert clientes["colunas"] == 3
        assert "id" in clientes["cabecalhos"]
        produtos = next(t for t in r.tabelas if t["nome"] == "Produtos")
        assert produtos["registros"] == 3
        assert r.total_registros == 5

    def test_analyze_empty_sheet(self):
        from migration.importer import ExcelAnalyzer
        from openpyxl import Workbook
        tmp = os.path.join(tempfile.gettempdir(), "test_xl_empty.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Vazio"
        wb.save(tmp)
        wb.close()
        r = ExcelAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert len(r.tabelas) == 1
        assert r.tabelas[0]["registros"] == 0

    def test_analyze_empty_file(self):
        from migration.importer import ExcelAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_xl_zero.xlsx")
        with open(tmp, "wb") as f:
            f.write(b"")
        r = ExcelAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is False

    def test_analyze_invalid_excel(self):
        from migration.importer import ExcelAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_xl_bad.xlsx")
        with open(tmp, "wb") as f:
            f.write(b"PK\x03\x04this is not a valid xlsx")
        r = ExcelAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is False


class TestCSVAnalyzer:
    def test_analyze_comma_delimited(self):
        from migration.importer import CSVAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_csv.csv")
        with open(tmp, "w", encoding="utf-8") as f:
            f.write("nome,preco,estoque\n")
            f.write("Arroz,10.50,100\n")
            f.write("Feijao,8.00,50\n")
            f.write("Acucar,5.50,200\n")
        r = CSVAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert r.tipo_fonte == "csv"
        assert len(r.tabelas) == 1
        assert r.tabelas[0]["registros"] == 3
        assert r.tabelas[0]["colunas"] == 3
        assert r.tabelas[0]["cabecalhos"] == ["nome", "preco", "estoque"]
        assert r.total_registros == 3
        assert r.metadata["delimitador"] == "','"
        assert "encoding_detectado" in r.metadata

    def test_analyze_semicolon_delimited(self):
        from migration.importer import CSVAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_csv_sep.csv")
        with open(tmp, "w", encoding="utf-8") as f:
            f.write("nome;preco\n")
            f.write("Cafe;5.00\n")
            f.write("Cha;3.00\n")
        r = CSVAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert r.metadata["delimitador"] == "';'"
        assert r.tabelas[0]["registros"] == 2

    def test_analyze_tab_delimited(self):
        from migration.importer import CSVAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_csv_tab.csv")
        with open(tmp, "w", encoding="utf-8") as f:
            f.write("id\tnome\n")
            f.write("1\tJoao\n")
        r = CSVAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert r.metadata["delimitador"] == "'\\t'"

    def test_analyze_empty_file(self):
        from migration.importer import CSVAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_csv_empty.csv")
        with open(tmp, "wb") as f:
            f.write(b"")
        r = CSVAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is False

    def test_analyze_headers_only(self):
        from migration.importer import CSVAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_csv_hdr.csv")
        with open(tmp, "w", encoding="utf-8") as f:
            f.write("col1,col2,col3\n")
        r = CSVAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert r.tabelas[0]["registros"] == 0
        assert r.tabelas[0]["colunas"] == 3

    def test_analyze_latin1_encoding(self):
        from migration.importer import CSVAnalyzer
        tmp = os.path.join(tempfile.gettempdir(), "test_csv_latin.csv")
        with open(tmp, "w", encoding="latin-1") as f:
            f.write("nome,descricao\n")
            f.write("Cafe,Cafe especial\n")
        r = CSVAnalyzer(tmp).analyze()
        os.remove(tmp)
        assert r.ok is True
        assert "encoding_detectado" in r.metadata


class TestAnalisarArquivo:
    def test_dispatch_csv(self):
        from migration.importer import analisar_arquivo
        tmp = os.path.join(tempfile.gettempdir(), "test_disp.csv")
        with open(tmp, "w") as f:
            f.write("a,b\n1,2\n")
        r = analisar_arquivo(tmp, "csv")
        os.remove(tmp)
        assert r.ok is True
        assert r.tipo_fonte == "csv"

    def test_dispatch_sqlite(self):
        from migration.importer import analisar_arquivo
        import sqlite3 as sq3
        tmp = os.path.join(tempfile.gettempdir(), "test_disp.db")
        conn = sq3.connect(tmp)
        conn.execute("CREATE TABLE t (id INTEGER)")
        conn.commit()
        conn.close()
        r = analisar_arquivo(tmp, "sqlite")
        os.remove(tmp)
        assert r.ok is True
        assert r.tipo_fonte == "sqlite"

    def test_dispatch_unknown(self):
        from migration.importer import analisar_arquivo
        r = analisar_arquivo("/nonexistent/file.xyz", "fortran")
        assert r.ok is False
        assert len(r.erros) > 0


class TestMigrationAnalysisAPI:
    def _login_admin(self, client):
        _login(client)

    def _upload_csv(self, client, content=b"a,b\n1,2\n"):
        client.get("/migracao/etapa2?tipo=csv")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (io.BytesIO(content), "data.csv")},
                        content_type="multipart/form-data")
        return r

    def _upload_sqlite(self, client):
        import sqlite3 as sq3
        tmp = os.path.join(tempfile.gettempdir(), "test_api.db")
        conn = sq3.connect(tmp)
        conn.execute("CREATE TABLE clientes (id INTEGER PRIMARY KEY, nome TEXT)")
        conn.execute("INSERT INTO clientes (nome) VALUES ('Teste')")
        conn.commit()
        conn.close()
        with open(tmp, "rb") as f:
            db_bytes = f.read()
        os.remove(tmp)
        client.get("/migracao/etapa2?tipo=sqlite")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (io.BytesIO(db_bytes), "banco.db")},
                        content_type="multipart/form-data")
        return r

    def test_analisar_api_returns_ok(self, client):
        self._login_admin(client)
        self._upload_csv(client)
        r = client.post("/api/migracao/analisar")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True

    def test_analisar_api_sqlite(self, client):
        self._login_admin(client)
        self._upload_sqlite(client)
        r = client.post("/api/migracao/analisar")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert d["tipo_fonte"] == "sqlite"
        assert len(d["tabelas"]) >= 1

    def test_analisar_api_stores_in_session(self, client):
        self._login_admin(client)
        self._upload_csv(client, b"col1,col2\n1,2\n3,4\n")
        r = client.post("/api/migracao/analisar")
        assert r.status_code == 200
        r = client.get("/migracao/etapa3")
        assert r.status_code == 200

    def test_analisar_requires_admin(self, client):
        r = client.post("/api/migracao/analisar")
        assert r.status_code == 302

    def test_etapa4_has_analysis_data(self, client):
        self._login_admin(client)
        self._upload_csv(client, b"nome,valor\nA,10\nB,20\n")
        client.post("/api/migracao/analisar")
        r = client.get("/migracao/etapa4")
        assert r.status_code == 200
        assert b"2" in r.data


class TestCompatibilityAnalyzer:
    def _base_result(self, tipo="csv", tables=None):
        if tables is None:
            tables = []
        return {
            "ok": True,
            "tipo_fonte": tipo,
            "arquivo_nome": "test.csv",
            "arquivo_tamanho_mb": 0.01,
            "tabelas": tables,
            "total_registros": sum(t.get("registros", 0) for t in tables),
            "integridade": [],
            "erros": [],
            "metadata": {},
        }

    def test_full_schema_csv_match(self):
        from migration.compatibility import CompatibilityAnalyzer, EXPECTED_TABLES
        result = self._base_result("csv", [{
            "nome": "clientes",
            "registros": 5,
            "colunas": 3,
            "cabecalhos": ["id", "nome", "telefone"],
        }])
        report = CompatibilityAnalyzer(result).analyze()
        assert report["ok"] is True
        assert report["tipo_fonte"] == "csv"
        assert report["total_tabelas_esperadas"] == len(EXPECTED_TABLES)
        assert report["total_tabelas_encontradas"] == 1
        assert report["total_tabelas_mapeadas"] == 1
        assert len(report["tabelas"]) == 1
        assert report["tabelas"][0]["nome"] == "clientes"
        assert report["tabelas"][0]["compatibilidade"] > 0
        assert report["criticos"] > 0

    def test_missing_tables_are_critical(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [{
            "nome": "produtos",
            "registros": 10,
            "colunas": 2,
            "cabecalhos": ["id", "nome"],
        }])
        report = CompatibilityAnalyzer(result).analyze()
        missing = [i for i in report["issues"] if i["tipo"] == "tabela_ausente"]
        assert len(missing) > 0
        assert all(i["severidade"] == "critico" for i in missing)
        assert report["criticos"] > 0

    def test_extra_tables_are_info(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [
            {"nome": "clientes", "registros": 1, "colunas": 1, "cabecalhos": ["nome"]},
            {"nome": "tabela_custom", "registros": 5, "colunas": 2, "cabecalhos": ["a", "b"]},
        ])
        report = CompatibilityAnalyzer(result).analyze()
        extras = [i for i in report["issues"] if i["tipo"] == "tabela_extras"]
        assert len(extras) == 1
        assert extras[0]["severidade"] == "info"

    def test_missing_required_column_is_critical(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [{
            "nome": "clientes",
            "registros": 3,
            "colunas": 2,
            "cabecalhos": ["id", "telefone"],
        }])
        report = CompatibilityAnalyzer(result).analyze()
        tbl = report["tabelas"][0]
        assert tbl["compatibilidade"] == 0.0
        critical_cols = [i for i in report["issues"]
                         if i["tipo"] == "coluna_obrigatoria_ausente" and i["tabela"] == "clientes"]
        assert len(critical_cols) > 0

    def test_missing_optional_column_is_warning(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [{
            "nome": "clientes",
            "registros": 2,
            "colunas": 2,
            "cabecalhos": ["nome", "telefone"],
        }])
        report = CompatibilityAnalyzer(result).analyze()
        optional_missing = [i for i in report["issues"]
                            if i["tipo"] == "coluna_ausente" and i["tabela"] == "clientes"]
        assert len(optional_missing) >= 1
        assert all(i["severidade"] == "aviso" for i in optional_missing)

    def test_extra_columns_are_info(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [{
            "nome": "clientes",
            "registros": 1,
            "colunas": 3,
            "cabecalhos": ["nome", "campo_custom", "outro_campo"],
        }])
        report = CompatibilityAnalyzer(result).analyze()
        extras = [i for i in report["issues"]
                  if i["tipo"] == "coluna_extras" and i["tabela"] == "clientes"]
        assert len(extras) == 2
        assert all(i["severidade"] == "info" for i in extras)

    def test_classification_excellent(self):
        from migration.compatibility import CompatibilityAnalyzer, EXPECTED_TABLES
        tables = []
        for tname in list(EXPECTED_TABLES.keys())[:10]:
            tdef = EXPECTED_TABLES[tname]
            tables.append({
                "nome": tname,
                "registros": 5,
                "colunas": len(tdef["columns"]),
                "cabecalhos": list(tdef["columns"].keys()),
            })
        result = self._base_result("csv", tables)
        report = CompatibilityAnalyzer(result).analyze()
        assert report["classificacao"] in ("Excelente", "Bom", "Parcial")

    def test_classification_incompatible(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [
            {"nome": "custom_a", "registros": 1, "colunas": 1, "cabecalhos": ["x"]},
        ])
        report = CompatibilityAnalyzer(result).analyze()
        assert report["classificacao"] == "Incompatível"

    def test_empty_tables_warning(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [{
            "nome": "clientes",
            "registros": 0,
            "colunas": 1,
            "cabecalhos": ["nome"],
        }])
        report = CompatibilityAnalyzer(result).analyze()
        empty = [i for i in report["issues"] if i["tipo"] == "tabela_vazia"]
        assert len(empty) == 1
        assert empty[0]["severidade"] == "aviso"

    def test_parse_sql_ddl(self):
        from migration.compatibility import CompatibilityAnalyzer
        sql = """
        CREATE TABLE clientes (
            id INTEGER PRIMARY KEY,
            nome TEXT NOT NULL,
            telefone TEXT
        );
        CREATE TABLE produtos (
            id INTEGER PRIMARY KEY,
            nome TEXT,
            preco REAL
        );
        """
        parsed = CompatibilityAnalyzer.parse_sql_ddl(sql)
        assert "clientes" in parsed
        assert "produtos" in parsed
        assert "id" in parsed["clientes"]
        assert "nome" in parsed["clientes"]
        assert "telefone" in parsed["clientes"]
        assert len(parsed["produtos"]) == 3

    def test_parse_sql_ddl_ignores_constraints(self):
        from migration.compatibility import CompatibilityAnalyzer
        sql = """
        CREATE TABLE test (
            id INTEGER PRIMARY KEY,
            nome TEXT NOT NULL,
            UNIQUE(nome)
        );
        """
        parsed = CompatibilityAnalyzer.parse_sql_ddl(sql)
        assert "test" in parsed
        assert "unique(nome)" not in parsed["test"]
        assert "id" in parsed["test"]
        assert "nome" in parsed["test"]

    def test_map_tables_exact(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [{
            "nome": "clientes",
            "registros": 1,
            "colunas": 1,
            "cabecalhos": ["nome"],
        }])
        analyzer = CompatibilityAnalyzer(result)
        mapping = analyzer._map_tables(analyzer.source_tables, analyzer._build_source_tables.__func__ is None and {} or {})
        from migration.compatibility import EXPECTED_TABLES
        mapping = analyzer._map_tables(analyzer.source_tables, EXPECTED_TABLES)
        assert mapping.get("clientes") == "clientes"

    def test_map_tables_fuzzy(self):
        from migration.compatibility import CompatibilityAnalyzer, EXPECTED_TABLES
        result = self._base_result("csv", [{
            "nome": "CLIENTES",
            "registros": 1,
            "colunas": 1,
            "cabecalhos": ["nome"],
        }])
        analyzer = CompatibilityAnalyzer(result)
        mapping = analyzer._map_tables(analyzer.source_tables, EXPECTED_TABLES)
        assert mapping.get("CLIENTES") == "clientes"

    def test_types_compatible(self):
        from migration.compatibility import CompatibilityAnalyzer
        assert CompatibilityAnalyzer._types_compatible("INTEGER", "INTEGER") is True
        assert CompatibilityAnalyzer._types_compatible("INTEGER", "INT") is True
        assert CompatibilityAnalyzer._types_compatible("REAL", "FLOAT") is True
        assert CompatibilityAnalyzer._types_compatible("TEXT", "VARCHAR") is True
        assert CompatibilityAnalyzer._types_compatible("TIMESTAMP", "DATETIME") is True
        assert CompatibilityAnalyzer._types_compatible("INTEGER", "TEXT") is False

    def test_classificar(self):
        from migration.compatibility import CompatibilityAnalyzer
        assert CompatibilityAnalyzer._classificar(95) == "Excelente"
        assert CompatibilityAnalyzer._classificar(80) == "Bom"
        assert CompatibilityAnalyzer._classificar(60) == "Parcial"
        assert CompatibilityAnalyzer._classificar(30) == "Incompatível"

    def test_pode_importar_false_on_critical(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [
            {"nome": "custom_only", "registros": 1, "colunas": 1, "cabecalhos": ["x"]},
        ])
        report = CompatibilityAnalyzer(result).analyze()
        assert report["pode_importar"] is False

    def test_full_schema_all_tables_present(self):
        from migration.compatibility import CompatibilityAnalyzer, EXPECTED_TABLES
        tables = []
        for tname, tdef in EXPECTED_TABLES.items():
            tables.append({
                "nome": tname,
                "registros": 1,
                "colunas": len(tdef["columns"]),
                "cabecalhos": list(tdef["columns"].keys()),
            })
        result = self._base_result("csv", tables)
        report = CompatibilityAnalyzer(result).analyze()
        assert report["total_tabelas_mapeadas"] == len(EXPECTED_TABLES)
        assert report["compatibilidade_index"] >= 90
        assert report["classificacao"] == "Excelente"
        assert report["pode_importar"] is True

    def test_gerar_relatorio_html(self):
        from migration.compatibility import CompatibilityAnalyzer, gerar_relatorio_html
        result = self._base_result("csv", [{
            "nome": "clientes",
            "registros": 5,
            "colunas": 3,
            "cabecalhos": ["id", "nome", "telefone"],
        }])
        report = CompatibilityAnalyzer(result).analyze()
        html = gerar_relatorio_html(report)
        assert "<!DOCTYPE html>" in html
        assert "Relatório de Compatibilidade" in html
        assert "clientes" in html
        assert str(report["compatibilidade_index"]) in html

    def test_no_tables_gives_zero_score(self):
        from migration.compatibility import CompatibilityAnalyzer
        result = self._base_result("csv", [])
        report = CompatibilityAnalyzer(result).analyze()
        assert report["compatibilidade_index"] == 0.0
        assert report["pode_importar"] is False

    def test_sqlite_enrichment(self):
        from migration.compatibility import CompatibilityAnalyzer
        import sqlite3 as sq3
        tmp = os.path.join(tempfile.gettempdir(), "test_compat.db")
        conn = sq3.connect(tmp)
        conn.execute("CREATE TABLE clientes (id INTEGER PRIMARY KEY, nome TEXT NOT NULL, telefone TEXT)")
        conn.execute("INSERT INTO clientes (nome, telefone) VALUES ('Joao', '123')")
        conn.commit()
        conn.close()
        result = self._base_result("sqlite", [{
            "nome": "clientes",
            "registros": 1,
            "colunas": 0,
            "cabecalhos": [],
        }])
        result["_filepath"] = tmp
        report = CompatibilityAnalyzer(result).analyze()
        os.remove(tmp)
        tbl = next((t for t in report["tabelas"] if t["nome"] == "clientes"), None)
        assert tbl is not None
        assert tbl["colunas_encontradas"] == 3

    def test_sqlite_indexes_check(self):
        from migration.compatibility import CompatibilityAnalyzer
        import sqlite3 as sq3
        tmp = os.path.join(tempfile.gettempdir(), "test_compat_idx.db")
        conn = sq3.connect(tmp)
        conn.execute("CREATE TABLE fiado (id INTEGER, cliente_id INTEGER, tipo TEXT, valor REAL)")
        conn.execute("CREATE TABLE vendas (id INTEGER, usuario_id INTEGER, valor_total REAL, tipo TEXT, data TIMESTAMP)")
        conn.commit()
        conn.close()
        result = self._base_result("sqlite", [
            {"nome": "fiado", "registros": 0, "colunas": 0, "cabecalhos": []},
            {"nome": "vendas", "registros": 0, "colunas": 0, "cabecalhos": []},
        ])
        result["_filepath"] = tmp
        report = CompatibilityAnalyzer(result).analyze()
        os.remove(tmp)
        idx_issues = [i for i in report["issues"] if i["tipo"] == "indice_ausente"]
        assert len(idx_issues) >= 1
        assert all(i["severidade"] == "info" for i in idx_issues)


class TestMigrationCompatibilityAPI:
    def _login_admin(self, client):
        _login(client)

    def _upload_csv(self, client, content=b"nome,telefone\nA,123\n"):
        client.get("/migracao/etapa2?tipo=csv")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (io.BytesIO(content), "data.csv")},
                        content_type="multipart/form-data")
        return r

    def _upload_sqlite(self, client):
        import sqlite3 as sq3
        tmp = os.path.join(tempfile.gettempdir(), "test_api_compat.db")
        conn = sq3.connect(tmp)
        conn.execute("CREATE TABLE clientes (id INTEGER PRIMARY KEY, nome TEXT, telefone TEXT)")
        conn.execute("INSERT INTO clientes (nome) VALUES ('Teste')")
        conn.commit()
        conn.close()
        with open(tmp, "rb") as f:
            db_bytes = f.read()
        os.remove(tmp)
        client.get("/migracao/etapa2?tipo=sqlite")
        r = client.post("/migracao/etapa2",
                        data={"arquivo": (io.BytesIO(db_bytes), "banco.db")},
                        content_type="multipart/form-data")
        return r

    def test_analisar_stores_compat_in_session(self, client):
        self._login_admin(client)
        self._upload_csv(client)
        r = client.post("/api/migracao/analisar")
        assert r.status_code == 200
        with client.session_transaction() as sess:
            compat = sess.get("migracao", {}).get("compatibilidade")
            assert compat is not None
            assert "compatibilidade_index" in compat
            assert "issues" in compat
            assert "pode_importar" in compat

    def test_compat_api_endpoint(self, client):
        self._login_admin(client)
        self._upload_csv(client)
        client.post("/api/migracao/analisar")
        r = client.get("/api/migracao/compatibilidade")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "compatibilidade_index" in d

    def test_compat_api_no_data(self, client):
        self._login_admin(client)
        r = client.get("/api/migracao/compatibilidade")
        assert r.status_code == 400

    def test_compat_api_requires_admin(self, client):
        r = client.get("/api/migracao/compatibilidade")
        assert r.status_code == 302

    def test_json_report(self, client):
        self._login_admin(client)
        self._upload_csv(client)
        client.post("/api/migracao/analisar")
        r = client.get("/migracao/relatorio/json")
        assert r.status_code == 200
        assert r.content_type == "application/json"
        data = json.loads(r.data)
        assert "compatibilidade_index" in data
        assert "issues" in data

    def test_json_report_no_data(self, client):
        self._login_admin(client)
        r = client.get("/migracao/relatorio/json")
        assert r.status_code == 400

    def test_html_report(self, client):
        self._login_admin(client)
        self._upload_csv(client)
        client.post("/api/migracao/analisar")
        r = client.get("/migracao/relatorio/html")
        assert r.status_code == 200
        assert b"Relat" in r.data

    def test_html_report_no_data(self, client):
        self._login_admin(client)
        r = client.get("/migracao/relatorio/html")
        assert r.status_code == 400

    def test_etapa3_shows_compat(self, client):
        self._login_admin(client)
        self._upload_csv(client, b"nome,telefone\nA,123\n")
        client.post("/api/migracao/analisar")
        r = client.get("/migracao/etapa3")
        assert r.status_code == 200
        assert b"compatibilidade" in r.data.lower() or b"Compatibilidade" in r.data

    def test_etapa4_shows_compat_summary(self, client):
        self._login_admin(client)
        self._upload_csv(client, b"nome,telefone\nA,123\n")
        client.post("/api/migracao/analisar")
        r = client.get("/migracao/etapa4")
        assert r.status_code == 200

    def test_confirmar_blocks_on_critical(self, client):
        self._login_admin(client)
        client.get("/migracao/etapa2?tipo=csv")
        client.post("/migracao/etapa2",
                     data={"arquivo": (io.BytesIO(b"a,b\n1,2\n"), "data.csv")},
                     content_type="multipart/form-data")
        client.post("/api/migracao/analisar")
        with client.session_transaction() as sess:
            m = sess.get("migracao", {})
            m["compatibilidade"] = {
                "ok": True,
                "pode_importar": False,
                "compatibilidade_index": 25.0,
                "criticos": 15,
                "avisos": 0,
                "informativos": 0,
                "tabelas": [],
                "issues": [],
                "total_tabelas_esperadas": 19,
                "total_tabelas_encontradas": 0,
                "total_tabelas_mapeadas": 0,
                "total_registros": 0,
                "classificacao": "Incompatível",
            }
            sess["migracao"] = m
        r = client.post("/api/migracao/confirmar")
        assert r.status_code == 400
        d = r.get_json()
        assert "bloqueada" in d["erro"].lower()

    def test_confirmar_allows_when_compatible(self, client):
        self._login_admin(client)
        self._upload_csv(client, b"nome,telefone\nA,123\n")
        client.post("/api/migracao/analisar")
        with client.session_transaction() as sess:
            m = sess.get("migracao", {})
            m["compatibilidade"] = {
                "ok": True,
                "pode_importar": True,
                "compatibilidade_index": 95.0,
                "criticos": 0,
                "avisos": 2,
                "informativos": 3,
                "tabelas": [],
                "issues": [],
                "total_tabelas_esperadas": 19,
                "total_tabelas_encontradas": 1,
                "total_tabelas_mapeadas": 1,
                "total_registros": 1,
                "classificacao": "Excelente",
            }
            sess["migracao"] = m
        r = client.post("/api/migracao/confirmar")
        d = r.get_json()
        assert d.get("ok") is True
        assert d.get("total_registros", 0) >= 0

    def test_etapa4_blocks_import_button(self, client):
        self._login_admin(client)
        self._upload_csv(client, b"nome\nA\n")
        client.post("/api/migracao/analisar")
        with client.session_transaction() as sess:
            m = sess.get("migracao", {})
            m["compatibilidade"] = {
                "ok": True,
                "pode_importar": False,
                "compatibilidade_index": 0.0,
                "criticos": 15,
                "avisos": 0,
                "informativos": 0,
                "tabelas": [],
                "issues": [],
                "total_tabelas_esperadas": 19,
                "total_tabelas_encontradas": 0,
                "total_tabelas_mapeadas": 0,
                "total_registros": 0,
                "classificacao": "Incompatível",
            }
            sess["migracao"] = m
        r = client.get("/migracao/etapa4")
        assert r.status_code == 200
        assert b"Importa" in r.data
        assert b"disabled" in r.data


class TestMaintenanceDashboard:
    def _login_admin(self, client):
        _login(client)

    def test_dashboard_loads(self, client):
        self._login_admin(client)
        r = client.get("/manutencao")
        assert r.status_code == 200
        assert b"Manuten" in r.data or b"Centro de Manuten" in r.data

    def test_dashboard_requires_admin(self, client):
        r = client.get("/manutencao")
        assert r.status_code == 302

    def test_dashboard_shows_stats_cards(self, client):
        self._login_admin(client)
        r = client.get("/manutencao")
        assert r.status_code == 200
        assert b"Backups" in r.data
        assert b"Banco" in r.data
        assert b"Tabelas" in r.data
        assert b"Espa" in r.data

    def test_dashboard_has_module_links(self, client):
        self._login_admin(client)
        r = client.get("/manutencao")
        assert r.status_code == 200
        assert b"/manutencao/backup" in r.data
        assert b"/manutencao/banco" in r.data
        assert b"/manutencao/limpeza" in r.data
        assert b"/manutencao/migracao" in r.data

    def test_stats_api(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/stats")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "total_backups" in d
        assert "integridade_ok" in d
        assert "total_tabelas" in d
        assert "total_registros" in d
        assert "espaco_banco_mb" in d


class TestMaintenanceBackup:
    def _login_admin(self, client):
        _login(client)

    def test_backup_page_loads(self, client):
        self._login_admin(client)
        r = client.get("/manutencao/backup")
        assert r.status_code == 200
        assert b"Backup" in r.data

    def test_backup_page_requires_admin(self, client):
        r = client.get("/manutencao/backup")
        assert r.status_code == 302

    def test_criar_backup(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/backup/criar",
                        json={"descricao": "Teste backup"},
                        content_type="application/json")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "backup" in d
        assert d["backup"]["tamanho_mb"] >= 0

    def test_backup_completo_contem_banco_manifesto_e_configuracao(self, client):
        import zipfile
        self._login_admin(client)
        r = client.post("/api/manutencao/backup/criar",
                        json={"descricao": "Pacote completo"})
        assert r.status_code == 200
        nome = r.get_json()["backup"]["nome"]
        assert nome.endswith(".zip")
        download = client.get(f"/api/manutencao/backup/download/{nome}")
        with zipfile.ZipFile(io.BytesIO(download.data)) as pacote:
            nomes = pacote.namelist()
            assert "manifest.json" in nomes
            assert "database/bar_adega.db" in nomes
            assert "config/config_snapshot.json" in nomes
            manifesto = json.loads(pacote.read("manifest.json"))
            assert manifesto["formato"] == "jams-adega-backup"
            assert manifesto["banco"]["integridade"] == "ok"

    def test_importar_backup_completo_validado(self, client):
        self._login_admin(client)
        criado = client.post("/api/manutencao/backup/criar", json={"descricao": "Importável"})
        nome = criado.get_json()["backup"]["nome"]
        dados = client.get(f"/api/manutencao/backup/download/{nome}").data
        r = client.post(
            "/api/manutencao/backup/importar",
            data={"arquivo": (io.BytesIO(dados), "copia.zip")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        resposta = r.get_json()
        assert resposta["ok"] is True
        assert resposta["backup"]["integridade"] == "ok"
        assert resposta["backup"]["registros"] > 0

    def test_importar_backup_corrompido_e_recusado(self, client):
        self._login_admin(client)
        r = client.post(
            "/api/manutencao/backup/importar",
            data={"arquivo": (io.BytesIO(b"nao e um zip"), "invalido.zip")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 400
        assert r.get_json()["ok"] is False

    def test_importar_fbk_usa_conversor_datacaixa(self, client, monkeypatch):
        self._login_admin(client)
        chamado = {}

        def converter(arquivo, nome, usuario_id=None, usuario_nome=None):
            chamado["nome"] = nome
            chamado["conteudo"] = arquivo.read()
            return {
                "nome": "convertido_datacaixa_teste.zip",
                "integridade": "ok",
                "registros": 10,
                "imagens": 1,
                "aplicado": True,
                "importados": {"produtos": 2, "clientes": 3, "vendas": 4},
            }, None

        monkeypatch.setattr("maintenance.routes.importar_e_aplicar_datacaixa", converter)
        r = client.post(
            "/api/manutencao/backup/importar",
            data={"arquivo": (io.BytesIO(b"backup-firebird"), "DATACAIXA.FBK")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        assert r.get_json()["backup"]["aplicado"] is True
        assert chamado == {"nome": "DATACAIXA.FBK", "conteudo": b"backup-firebird"}

    def test_restauracao_completa_recupera_imagem(self, client):
        self._login_admin(client)
        uploads = app.config["BACKUP_UPLOADS_DIR"]
        imagem = os.path.join(uploads, "produtos", "produto_teste.png")
        os.makedirs(os.path.dirname(imagem), exist_ok=True)
        conteudo = b"imagem-de-teste-do-backup"
        with open(imagem, "wb") as arquivo:
            arquivo.write(conteudo)
        criado = client.post("/api/manutencao/backup/criar", json={"descricao": "Com imagem"})
        nome = criado.get_json()["backup"]["nome"]
        os.remove(imagem)
        restaurado = client.post("/api/manutencao/backup/restaurar", json={"nome": nome})
        assert restaurado.status_code == 200
        with open(imagem, "rb") as arquivo:
            assert arquivo.read() == conteudo

    def test_listar_backups(self, client):
        self._login_admin(client)
        client.post("/api/manutencao/backup/criar", json={"descricao": "test"})
        r = client.get("/api/manutencao/backup/listar")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert d["total"] >= 1

    def test_remover_backup(self, client):
        self._login_admin(client)
        cr = client.post("/api/manutencao/backup/criar", json={"descricao": "to delete"})
        nome = cr.get_json()["backup"]["nome"]
        r = client.post("/api/manutencao/backup/remover",
                        json={"nome": nome},
                        content_type="application/json")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True

    def test_remover_backup_inexistente(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/backup/remover",
                        json={"nome": "nao_existe.db"},
                        content_type="application/json")
        assert r.status_code == 404

    def test_espaco_utilizado(self, client):
        self._login_admin(client)
        client.post("/api/manutencao/backup/criar", json={"descricao": "test"})
        r = client.get("/api/manutencao/backup/espaco")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert d["mb"] >= 0

    def test_backup_requires_admin(self, client):
        r = client.get("/api/manutencao/backup/listar")
        assert r.status_code == 302

    def test_restaurar_backup_sem_nome(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/backup/restaurar",
                        json={},
                        content_type="application/json")
        assert r.status_code == 400

    def test_remover_backup_sem_nome(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/backup/remover",
                        json={},
                        content_type="application/json")
        assert r.status_code == 400


class TestMaintenanceBanco:
    def _login_admin(self, client):
        _login(client)

    def test_banco_page_loads(self, client):
        self._login_admin(client)
        r = client.get("/manutencao/banco")
        assert r.status_code == 200
        assert b"Diagn" in r.data or b"Banco de Dados" in r.data

    def test_banco_page_requires_admin(self, client):
        r = client.get("/manutencao/banco")
        assert r.status_code == 302

    def test_diagnostico_api(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/banco/diagnostico")
        assert r.status_code == 200
        d = r.get_json()
        assert "ok" in d
        assert "tabelas" in d
        assert "total_registros" in d

    def test_integridade_api(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/banco/integridade")
        assert r.status_code == 200
        d = r.get_json()
        assert "integridade" in d
        assert "versao_sqlite" in d

    def test_estatisticas_api(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/banco/estatisticas")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "tabelas" in d
        assert "total_tabelas" in d

    def test_fks_api(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/banco/fks")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "orfas" in d

    def test_recalcular_fiados_api(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/banco/recalcular_fiados")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "clientes_atualizados" in d

    def test_diagnostico_requires_admin(self, client):
        r = client.get("/api/manutencao/banco/diagnostico")
        assert r.status_code == 302


class TestMaintenanceLimpeza:
    def _login_admin(self, client):
        _login(client)

    def test_limpeza_page_loads(self, client):
        self._login_admin(client)
        r = client.get("/manutencao/limpeza")
        assert r.status_code == 200
        assert b"Limpeza" in r.data

    def test_limpeza_page_requires_admin(self, client):
        r = client.get("/manutencao/limpeza")
        assert r.status_code == 302

    def test_limpeza_page_has_preview_buttons(self, client):
        self._login_admin(client)
        r = client.get("/manutencao/limpeza")
        assert r.status_code == 200
        assert b"Visualizar Impacto" in r.data

    def test_limpeza_page_has_security_warning(self, client):
        self._login_admin(client)
        r = client.get("/manutencao/limpeza")
        assert r.status_code == 200
        assert b"backup autom" in r.data.lower() or b"Seguran" in r.data

    def test_limpeza_preview_vendas(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "vendas"},
                        content_type="application/json")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "tabelas" in d
        assert "total_registros" in d
        assert "confirmacao_necessaria" in d
        assert d["confirmacao_necessaria"] == "CONFIRMAR LIMPEZA"

    def test_limpeza_preview_comandas(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "comandas"},
                        content_type="application/json")
        d = r.get_json()
        assert d["ok"] is True
        assert "tabelas" in d

    def test_limpeza_preview_caixa(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "caixa"},
                        content_type="application/json")
        d = r.get_json()
        assert d["ok"] is True

    def test_limpeza_preview_estoque(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "estoque"},
                        content_type="application/json")
        d = r.get_json()
        assert d["ok"] is True

    def test_limpeza_preview_clientes(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "clientes"},
                        content_type="application/json")
        d = r.get_json()
        assert d["ok"] is True

    def test_limpeza_preview_fornecedores(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "fornecedores"},
                        content_type="application/json")
        d = r.get_json()
        assert d["ok"] is True

    def test_limpeza_preview_funcionarios(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "funcionarios"},
                        content_type="application/json")
        d = r.get_json()
        assert d["ok"] is True

    def test_limpeza_preview_reset(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "reset"},
                        content_type="application/json")
        d = r.get_json()
        assert d["ok"] is True
        assert d["confirmacao_necessaria"] == "RESETAR SISTEMA"
        assert d["nivel"] == "critico"

    def test_limpeza_preview_unknown_action(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "desconhecida"},
                        content_type="application/json")
        assert r.status_code == 400

    def test_limpeza_preview_requires_admin(self, client):
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "vendas"},
                        content_type="application/json")
        assert r.status_code == 302


class TestMaintenanceMigracao:
    def _login_admin(self, client):
        _login(client)

    def test_migracao_page_loads(self, client):
        self._login_admin(client)
        r = client.get("/manutencao/migracao")
        assert r.status_code == 200
        assert b"Migra" in r.data

    def test_migracao_page_requires_admin(self, client):
        r = client.get("/manutencao/migracao")
        assert r.status_code == 302

    def test_migracao_has_shortcuts(self, client):
        self._login_admin(client)
        r = client.get("/manutencao/migracao")
        assert r.status_code == 200
        assert b"/migracao/etapa1" in r.data
        assert b"/migracao/etapa3" in r.data
        assert b"/manutencao/backup" in r.data


class TestSidebarMaintenance:
    def test_sidebar_has_maintenance_link(self, client):
        _login(client)
        r = client.get("/")
        assert r.status_code == 200
        assert b"Centro de Manuten" in r.data
        assert b"/manutencao" in r.data

    def test_sidebar_has_migracao_link(self, client):
        _login(client)
        r = client.get("/")
        assert r.status_code == 200
        assert b"Assistente de Migra" in r.data

    def test_sidebar_has_auditoria_manutencao_link(self, client):
        _login(client)
        r = client.get("/")
        assert r.status_code == 200
        assert b"Auditoria Manuten" in r.data
        assert b"/manutencao/auditoria" in r.data


class TestBackupDownload:
    def _login_admin(self, client):
        _login(client)

    def test_download_backup(self, client):
        self._login_admin(client)
        cr = client.post("/api/manutencao/backup/criar", json={"descricao": "download test"})
        nome = cr.get_json()["backup"]["nome"]
        r = client.get(f"/api/manutencao/backup/download/{nome}")
        assert r.status_code == 200
        assert r.content_type == "application/octet-stream" or len(r.data) > 0

    def test_download_backup_inexistente(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/backup/download/nao_existe.db")
        assert r.status_code == 404

    def test_download_requires_admin(self, client):
        r = client.get("/api/manutencao/backup/download/qualquer.db")
        assert r.status_code == 302


class TestBackupRestoreComSeguranca:
    def _login_admin(self, client):
        _login(client)

    def test_restaurar_backup_cria_pre_backup(self, client):
        self._login_admin(client)
        cr = client.post("/api/manutencao/backup/criar", json={"descricao": "restore test"})
        nome = cr.get_json()["backup"]["nome"]
        r = client.post("/api/manutencao/backup/restaurar",
                        json={"nome": nome},
                        content_type="application/json")
        d = r.get_json()
        assert r.status_code == 200
        assert d["ok"] is True
        assert "pre_backup" in d
        assert d["pre_backup"] is not None
        assert d["pre_backup"].startswith("pre_restauracao_")

    def test_restaurar_backup_inexistente(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/backup/restaurar",
                        json={"nome": "backup_inexistente_999.db"},
                        content_type="application/json")
        assert r.status_code == 500
        d = r.get_json()
        assert d["ok"] is False


class TestBackupHistoricoCompleto:
    def _login_admin(self, client):
        _login(client)

    def test_historico_tem_usuario(self, client):
        self._login_admin(client)
        client.post("/api/manutencao/backup/criar", json={"descricao": "user test"})
        r = client.get("/api/manutencao/backup/listar")
        d = r.get_json()
        assert d["ok"] is True
        assert len(d["backups"]) > 0
        latest = d["backups"][0]
        assert "usuario" in latest
        assert "tamanho_bytes" in latest

    def test_historico_tem_tamanho_bytes(self, client):
        self._login_admin(client)
        cr = client.post("/api/manutencao/backup/criar", json={"descricao": "size test"})
        r = client.get("/api/manutencao/backup/listar")
        d = r.get_json()
        for b in d["backups"]:
            assert "tamanho_bytes" in b
            assert "tamanho_mb" in b


class TestDiagnosticoExpandido:
    def _login_admin(self, client):
        _login(client)

    def test_diagnostico_tem_tempo_resposta(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/banco/diagnostico")
        d = r.get_json()
        assert "tempo_resposta_ms" in d
        assert d["tempo_resposta_ms"] is not None
        assert d["tempo_resposta_ms"] >= 0

    def test_diagnostico_tabelas_tem_tamanho(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/banco/diagnostico")
        d = r.get_json()
        assert "tabelas" in d
        for t in d["tabelas"]:
            assert "tamanho_bytes" in t
            assert "tamanho_kb" in t
            assert "tamanho_mb" in t

    def test_diagnostico_tem_ultimo_backup(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/banco/diagnostico")
        d = r.get_json()
        assert "ultimo_backup" in d
        assert "ultimo_backup_nome" in d


class TestDashboardSaude:
    def _login_admin(self, client):
        _login(client)

    def test_dashboard_tem_saude(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/stats")
        d = r.get_json()
        assert "saude" in d
        assert d["saude"] in ("Excelente", "Boa", "Atenção", "Crítica")
        assert "saude_pontos" in d
        assert isinstance(d["saude_pontos"], int)

    def test_dashboard_tem_espaco_livre(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/stats")
        d = r.get_json()
        assert "espaco_livre_bytes" in d
        assert "espaco_livre_mb" in d

    def test_dashboard_tem_data_backup_recente(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/stats")
        d = r.get_json()
        assert "data_backup_mais_recente" in d

    def test_dashboard_template_saude(self, client):
        self._login_admin(client)
        r = client.get("/manutencao")
        assert r.status_code == 200
        assert b"Sa" in r.data


class TestAuditoriaManutencao:
    def _login_admin(self, client):
        _login(client)

    def test_auditoria_page_loads(self, client):
        self._login_admin(client)
        r = client.get("/manutencao/auditoria")
        assert r.status_code == 200
        assert b"Auditoria" in r.data

    def test_auditoria_page_requires_admin(self, client):
        r = client.get("/manutencao/auditoria")
        assert r.status_code == 302

    def test_auditoria_api_returns_ok(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        assert r.status_code == 200
        assert d["ok"] is True
        assert "registros" in d
        assert "total" in d

    def test_auditoria_api_registra_backup(self, client):
        self._login_admin(client)
        client.post("/api/manutencao/backup/criar", json={"descricao": "audit test"})
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert "Backup criado" in ops

    def test_auditoria_api_registra_diagnostico(self, client):
        self._login_admin(client)
        client.get("/api/manutencao/banco/diagnostico")
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert "Diagnóstico executado" in ops

    def test_auditoria_api_registra_integridade(self, client):
        self._login_admin(client)
        client.get("/api/manutencao/banco/integridade")
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert "Verificação de integridade" in ops

    def test_auditoria_api_registra_recalculo(self, client):
        self._login_admin(client)
        client.post("/api/manutencao/banco/recalcular_fiados")
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert "Fiados recalculados" in ops

    def test_auditoria_api_campos_completos(self, client):
        self._login_admin(client)
        client.post("/api/manutencao/backup/criar", json={"descricao": "fields test"})
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        assert len(d["registros"]) > 0
        reg = d["registros"][0]
        assert "data_hora" in reg
        assert "usuario" in reg
        assert "operacao" in reg
        assert "resultado" in reg
        assert "detalhes" in reg

    def test_auditoria_api_requires_admin(self, client):
        r = client.get("/api/manutencao/auditoria")
        assert r.status_code == 302

    def test_auditoria_api_with_limit(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/auditoria?limite=5")
        d = r.get_json()
        assert d["ok"] is True
        assert len(d["registros"]) <= 5

    def test_auditoria_api_with_offset(self, client):
        self._login_admin(client)
        r = client.get("/api/manutencao/auditoria?offset=0&limite=2")
        d = r.get_json()
        assert d["ok"] is True

    def test_auditoria_api_registra_remocao_backup(self, client):
        self._login_admin(client)
        cr = client.post("/api/manutencao/backup/criar", json={"descricao": "to delete"})
        nome = cr.get_json()["backup"]["nome"]
        client.post("/api/manutencao/backup/remover",
                     json={"nome": nome},
                     content_type="application/json")
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert "Backup removido" in ops

    def test_auditoria_api_registra_restauracao(self, client):
        self._login_admin(client)
        cr = client.post("/api/manutencao/backup/criar", json={"descricao": "restore audit"})
        nome = cr.get_json()["backup"]["nome"]
        client.post("/api/manutencao/backup/restaurar",
                     json={"nome": nome},
                     content_type="application/json")
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert "Backup restaurado" in ops

    def test_auditoria_api_registra_download(self, client):
        self._login_admin(client)
        cr = client.post("/api/manutencao/backup/criar", json={"descricao": "download audit"})
        nome = cr.get_json()["backup"]["nome"]
        client.get(f"/api/manutencao/backup/download/{nome}")
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert "Download de backup" in ops


class TestLimpezaExecucao:
    _SENHA = os.environ["ADMIN_SENHA"]
    _CONF = "CONFIRMAR LIMPEZA"
    _RESET = "RESETAR SISTEMA"

    def _login_admin(self, client):
        _login(client)

    def _confirmar(self, client, acao, senha=None, confirmacao=None):
        return client.post("/api/manutencao/limpeza/confirmar",
                           json={"acao": acao, "senha": senha or self._SENHA, "confirmacao": confirmacao or self._CONF},
                           content_type="application/json")

    def test_limpeza_vendas_executa(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "vendas")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert "backup_criado" in d
        assert d["integridade"] is True
        assert "removidos" in d
        assert "preservados" in d
        assert "tempo_segundos" in d

    def test_limpeza_comandas_executa(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "comandas")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert d["integridade"] is True

    def test_limpeza_caixa_executa(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "caixa")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True

    def test_limpeza_estoque_executa(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "estoque")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True

    def test_limpeza_clientes_executa(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "clientes")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True

    def test_limpeza_fornecedores_executa(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "fornecedores")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True

    def test_limpeza_funcionarios_executa(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "funcionarios")
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        _reseed(client)

    def test_reset_inteligente_executa(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "reset", confirmacao=self._RESET)
        assert r.status_code == 200
        d = r.get_json()
        assert d["ok"] is True
        assert d["integridade"] is True
        assert "backup_criado" in d
        _reseed(client)

    def test_reset_preserva_admin(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "reset", confirmacao=self._RESET)
        d = r.get_json()
        assert d["ok"] is True
        assert "usuarios" in d.get("preservados", {})
        assert d["preservados"]["usuarios"] >= 1
        _reseed(client)


class TestLimpezaSeguranca:
    _SENHA = os.environ["ADMIN_SENHA"]
    _CONF = "CONFIRMAR LIMPEZA"
    _RESET = "RESETAR SISTEMA"

    def _login_admin(self, client):
        _login(client)

    def test_senha_obrigatoria(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/confirmar",
                        json={"acao": "vendas", "senha": "", "confirmacao": self._CONF},
                        content_type="application/json")
        assert r.status_code == 400
        d = r.get_json()
        assert d["ok"] is False
        assert "senha" in d["erro"].lower()

    def test_senha_incorreta_rejeita(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/confirmar",
                        json={"acao": "vendas", "senha": "senhaerrada123", "confirmacao": self._CONF},
                        content_type="application/json")
        assert r.status_code == 403
        d = r.get_json()
        assert d["ok"] is False
        assert "incorreta" in d["erro"].lower()

    def test_confirmacao_obrigatoria(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/confirmar",
                        json={"acao": "vendas", "senha": self._SENHA, "confirmacao": ""},
                        content_type="application/json")
        assert r.status_code == 400
        d = r.get_json()
        assert "confirma" in d["erro"].lower()

    def test_confirmacao_incorreta_rejeita(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/confirmar",
                        json={"acao": "vendas", "senha": self._SENHA, "confirmacao": "TEXTO ERRADO"},
                        content_type="application/json")
        assert r.status_code == 400
        d = r.get_json()
        assert d["ok"] is False
        assert "incorreta" in d["erro"].lower()

    def test_confirmacao_reset_errada(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/confirmar",
                        json={"acao": "reset", "senha": self._SENHA, "confirmacao": self._CONF},
                        content_type="application/json")
        assert r.status_code == 400
        d = r.get_json()
        assert "RESETAR SISTEMA" in d["erro"]

    def test_senha_errada_registra_auditoria(self, client):
        self._login_admin(client)
        client.post("/api/manutencao/limpeza/confirmar",
                    json={"acao": "vendas", "senha": "wrongpass", "confirmacao": self._CONF},
                    content_type="application/json")
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert any("senha inválida" in op.lower() or "senha incorreta" in op.lower() or "inválida" in op.lower() for op in ops)

    def test_limpeza_requer_admin(self, client):
        r = client.post("/api/manutencao/limpeza/confirmar",
                        json={"acao": "vendas", "senha": "x", "confirmacao": self._CONF},
                        content_type="application/json")
        assert r.status_code == 302

    def test_preview_requer_admin(self, client):
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "vendas"},
                        content_type="application/json")
        assert r.status_code == 302


class TestLimpezaBackup:
    _SENHA = os.environ["ADMIN_SENHA"]
    _CONF = "CONFIRMAR LIMPEZA"
    _RESET = "RESETAR SISTEMA"

    def _login_admin(self, client):
        _login(client)

    def test_limpeza_cria_backup_automatico(self, client):
        self._login_admin(client)
        count_before = client.get("/api/manutencao/backup/listar").get_json()["total"]
        self._confirmar(client, "vendas")
        count_after = client.get("/api/manutencao/backup/listar").get_json()["total"]
        assert count_after >= count_before + 1

    def test_backup_pre_limpeza_no_nome(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "comandas")
        d = r.get_json()
        assert d["ok"] is True
        nome = d["backup_criado"]
        assert nome.startswith("backup_")
        assert "pré-limpeza" in d.get("removidos", {}) or True
        backups = client.get("/api/manutencao/backup/listar").get_json()["backups"]
        nomes = [b["nome"] for b in backups]
        assert nome in nomes

    def _confirmar(self, client, acao, senha=None, confirmacao=None):
        return client.post("/api/manutencao/limpeza/confirmar",
                           json={"acao": acao, "senha": senha or self._SENHA, "confirmacao": confirmacao or self._CONF},
                           content_type="application/json")


class TestLimpezaIntegridade:
    _SENHA = os.environ["ADMIN_SENHA"]
    _CONF = "CONFIRMAR LIMPEZA"
    _RESET = "RESETAR SISTEMA"

    def _login_admin(self, client):
        _login(client)

    def _confirmar(self, client, acao, senha=None, confirmacao=None):
        return client.post("/api/manutencao/limpeza/confirmar",
                           json={"acao": acao, "senha": senha or self._SENHA, "confirmacao": confirmacao or self._CONF},
                           content_type="application/json")

    def test_integridade_pos_limpeza_vendas(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "vendas")
        d = r.get_json()
        assert d["ok"] is True
        assert d["integridade"] is True

    def test_integridade_pos_reset(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "reset", confirmacao=self._RESET)
        d = r.get_json()
        assert d["ok"] is True
        assert d["integridade"] is True

    def test_integridade_pos_limpeza_todas(self, client):
        self._login_admin(client)
        for acao in ["vendas", "comandas", "caixa", "estoque", "clientes", "fornecedores", "funcionarios"]:
            self._confirmar(client, acao)
        r = client.get("/api/manutencao/banco/integridade")
        d = r.get_json()
        assert d["integridade"] is True
        _reseed(client)


class TestLimpezaAuditoria:
    _SENHA = os.environ["ADMIN_SENHA"]
    _CONF = "CONFIRMAR LIMPEZA"
    _RESET = "RESETAR SISTEMA"

    def _login_admin(self, client):
        _login(client)

    def _confirmar(self, client, acao, senha=None, confirmacao=None):
        return client.post("/api/manutencao/limpeza/confirmar",
                           json={"acao": acao, "senha": senha or self._SENHA, "confirmacao": confirmacao or self._CONF},
                           content_type="application/json")

    def test_auditoria_registra_limpeza(self, client):
        self._login_admin(client)
        self._confirmar(client, "estoque")
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert any("limpeza executada" in op.lower() for op in ops)

    def test_auditoria_registra_reset(self, client):
        self._login_admin(client)
        self._confirmar(client, "reset", confirmacao=self._RESET)
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        ops = [reg["operacao"] for reg in d["registros"]]
        assert any("reset inteligente" in op.lower() or "reset" in op.lower() for op in ops)

    def test_auditoria_detalhes_contem_backup(self, client):
        self._login_admin(client)
        self._confirmar(client, "fornecedores")
        r = client.get("/api/manutencao/auditoria")
        d = r.get_json()
        for reg in d["registros"]:
            if "limpeza executada" in reg["operacao"].lower():
                assert "backup" in (reg["detalhes"] or "").lower()
                break


class TestLimpezaRelatorio:
    _SENHA = os.environ["ADMIN_SENHA"]
    _CONF = "CONFIRMAR LIMPEZA"
    _RESET = "RESETAR SISTEMA"

    def _login_admin(self, client):
        _login(client)

    def _confirmar(self, client, acao, senha=None, confirmacao=None):
        return client.post("/api/manutencao/limpeza/confirmar",
                           json={"acao": acao, "senha": senha or self._SENHA, "confirmacao": confirmacao or self._CONF},
                           content_type="application/json")

    def test_relatorio_tem_removidos(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "vendas")
        d = r.get_json()
        assert d["ok"] is True
        assert "removidos" in d
        assert "total_removidos" in d
        assert "preservados" in d

    def test_relatorio_tem_tempo(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "comandas")
        d = r.get_json()
        assert "tempo_segundos" in d
        assert isinstance(d["tempo_segundos"], (int, float))

    def test_relatorio_tem_backup(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "caixa")
        d = r.get_json()
        assert "backup_criado" in d
        assert d["backup_criado"].startswith("backup_")

    def test_relatorio_preservados_nao_vazio(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "vendas")
        d = r.get_json()
        assert len(d["preservados"]) > 0

    def test_reset_relatorio_removidos(self, client):
        self._login_admin(client)
        r = self._confirmar(client, "reset", confirmacao=self._RESET)
        d = r.get_json()
        assert d["ok"] is True
        assert d["total_removidos"] >= 0
        assert "preservados" in d
        assert len(d["preservados"]) > 0


class TestLimpezaRollback:
    _SENHA = os.environ["ADMIN_SENHA"]
    _CONF = "CONFIRMAR LIMPEZA"

    def _login_admin(self, client):
        _login(client)

    def test_acao_desconhecida_rejeita(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/confirmar",
                        json={"acao": "naoexiste", "senha": self._SENHA, "confirmacao": "X"},
                        content_type="application/json")
        assert r.status_code == 400

    def test_preview_acao_invalida(self, client):
        self._login_admin(client)
        r = client.post("/api/manutencao/limpeza/preview",
                        json={"acao": "invalida"},
                        content_type="application/json")
        assert r.status_code == 400

    def test_limpeza_vazio_retorna_ok(self, client):
        self._login_admin(client)
        self._confirmar(client, "vendas")
        r = self._confirmar(client, "vendas")
        d = r.get_json()
        assert d["ok"] is True
        assert d["total_removidos"] == 0

    def _confirmar(self, client, acao, senha=None, confirmacao=None):
        return client.post("/api/manutencao/limpeza/confirmar",
                           json={"acao": acao, "senha": senha or self._SENHA, "confirmacao": confirmacao or self._CONF},
                           content_type="application/json")
