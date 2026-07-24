import os
import re
import csv
import time
import sqlite3
import chardet


class AnalysisResult:
    def __init__(self):
        self.ok = True
        self.tipo_fonte = ""
        self.arquivo_nome = ""
        self.arquivo_tamanho_mb = 0
        self.tabelas = []
        self.total_registros = 0
        self.integridade = []
        self.erros = []
        self.metadata = {}

    def to_dict(self):
        return {
            "ok": self.ok,
            "tipo_fonte": self.tipo_fonte,
            "arquivo_nome": self.arquivo_nome,
            "arquivo_tamanho_mb": self.arquivo_tamanho_mb,
            "tabelas": self.tabelas,
            "total_registros": self.total_registros,
            "integridade": self.integridade,
            "erros": self.erros,
            "metadata": self.metadata,
        }


class SQLiteAnalyzer:
    def __init__(self, filepath):
        self.filepath = filepath

    def analyze(self):
        result = AnalysisResult()
        result.tipo_fonte = "sqlite"
        result.arquivo_nome = os.path.basename(self.filepath)
        result.arquivo_tamanho_mb = round(os.path.getsize(self.filepath) / (1024 * 1024), 2)
        result._filepath = self.filepath

        if os.path.getsize(self.filepath) == 0:
            result.ok = False
            result.erros.append("Arquivo vazio")
            return result

        conn = None
        try:
            conn = sqlite3.connect(f"file:{self.filepath}?mode=ro", uri=True)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            try:
                ver = cur.execute("PRAGMA compile_options").fetchall()
                version_rows = [r[0] for r in ver if "VERSION" in r[0]]
                result.metadata["sqlite_version"] = version_rows[0].split("=")[-1] if version_rows else "desconhecida"
            except Exception:
                result.metadata["sqlite_version"] = "desconhecida"

            try:
                integrity = cur.execute("PRAGMA integrity_check").fetchone()
                is_ok = integrity[0] == "ok"
                result.integridade.append({
                    "check": "integrity_check",
                    "ok": is_ok,
                    "detalhe": "Banco íntegro" if is_ok else integrity[0],
                })
            except Exception as e:
                result.integridade.append({
                    "check": "integrity_check",
                    "ok": False,
                    "detalhe": f"Erro: {str(e)}",
                })

            tables = cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
            ).fetchall()

            for tbl in tables:
                tname = tbl[0]
                try:
                    count = cur.execute(f'SELECT COUNT(*) FROM "{tname}"').fetchone()[0]
                except Exception:
                    count = -1
                result.tabelas.append({"nome": tname, "registros": count})
                if count > 0:
                    result.total_registros += count

            result.integridade.append({
                "check": "tabelas",
                "ok": len(result.tabelas) > 0,
                "detalhe": f"{len(result.tabelas)} tabela(s) encontrada(s)",
            })

            result.integridade.append({
                "check": "estrutura",
                "ok": True,
                "detalhe": "Estrutura SQLite compatível",
            })

        except sqlite3.DatabaseError:
            result.ok = False
            result.erros.append("Arquivo não é um banco SQLite válido ou está corrompido")
        except Exception as e:
            result.ok = False
            result.erros.append(f"Erro ao analisar: {str(e)}")
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

        return result


class SQLAnalyzer:
    def __init__(self, filepath):
        self.filepath = filepath

    def analyze(self):
        result = AnalysisResult()
        result.tipo_fonte = "sql"
        result.arquivo_nome = os.path.basename(self.filepath)
        result.arquivo_tamanho_mb = round(os.path.getsize(self.filepath) / (1024 * 1024), 2)

        if os.path.getsize(self.filepath) == 0:
            result.ok = False
            result.erros.append("Arquivo vazio")
            return result

        try:
            raw = open(self.filepath, "rb").read()
            det = chardet.detect(raw)
            encoding = det.get("encoding", "utf-8") or "utf-8"
            result.metadata["encoding_detectado"] = encoding

            content = raw.decode(encoding, errors="replace")
            result.metadata["total_linhas"] = content.count("\n") + 1

            table_pattern = re.compile(
                r"CREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?[`\"']?(\w+)[`\"']?\s*\(",
                re.IGNORECASE,
            )
            tables_found = table_pattern.findall(content)
            result.metadata["tabelas_create"] = tables_found

            insert_pattern = re.compile(
                r"INSERT\s+INTO\s+[`\"']?(\w+)[`\"']?\s",
                re.IGNORECASE,
            )
            inserts = insert_pattern.findall(content)

            table_counts = {}
            for t in inserts:
                table_counts[t] = table_counts.get(t, 0) + 1

            for tname in tables_found:
                count = table_counts.get(tname, 0)
                result.tabelas.append({"nome": tname, "registros": count})
                result.total_registros += count

            for tname, count in table_counts.items():
                if tname not in tables_found:
                    result.tabelas.append({"nome": tname, "registros": count})
                    result.total_registros += count

            result.integridade.append({
                "check": "create_tables",
                "ok": len(tables_found) > 0,
                "detalhe": f"{len(tables_found)} CREATE TABLE(s) encontrado(s)",
            })
            result.integridade.append({
                "check": "inserts",
                "ok": len(inserts) > 0,
                "detalhe": f"{len(inserts)} INSERT INTO(s) encontrado(s)",
            })
            result.integridade.append({
                "check": "estrutura",
                "ok": True,
                "detalhe": "Arquivo SQL legível",
            })

        except Exception as e:
            result.ok = False
            result.erros.append(f"Erro ao analisar SQL: {str(e)}")

        return result


class ExcelAnalyzer:
    def __init__(self, filepath):
        self.filepath = filepath

    def analyze(self):
        result = AnalysisResult()
        result.tipo_fonte = "excel"
        result.arquivo_nome = os.path.basename(self.filepath)
        result.arquivo_tamanho_mb = round(os.path.getsize(self.filepath) / (1024 * 1024), 2)

        if os.path.getsize(self.filepath) == 0:
            result.ok = False
            result.erros.append("Arquivo vazio")
            return result

        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.filepath, read_only=True, data_only=True)
            result.metadata["abas"] = wb.sheetnames

            for sname in wb.sheetnames:
                ws = wb[sname]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    result.tabelas.append({"nome": sname, "registros": 0, "colunas": 0, "cabecalhos": []})
                    continue

                headers = [str(c) if c is not None else "" for c in rows[0]]
                data_rows = rows[1:]
                result.tabelas.append({
                    "nome": sname,
                    "registros": len(data_rows),
                    "colunas": len(headers),
                    "cabecalhos": headers,
                })
                result.total_registros += len(data_rows)

            result.integridade.append({
                "check": "abas",
                "ok": len(wb.sheetnames) > 0,
                "detalhe": f"{len(wb.sheetnames)} aba(s) encontrada(s)",
            })
            result.integridade.append({
                "check": "estrutura",
                "ok": True,
                "detalhe": "Planilha Excel legível",
            })
            wb.close()
        except ImportError:
            result.ok = False
            result.erros.append("Biblioteca openpyxl não instalada")
        except Exception as e:
            result.ok = False
            result.erros.append(f"Erro ao analisar Excel: {str(e)}")

        return result


class CSVAnalyzer:
    def __init__(self, filepath):
        self.filepath = filepath

    def analyze(self):
        result = AnalysisResult()
        result.tipo_fonte = "csv"
        result.arquivo_nome = os.path.basename(self.filepath)
        result.arquivo_tamanho_mb = round(os.path.getsize(self.filepath) / (1024 * 1024), 2)

        if os.path.getsize(self.filepath) == 0:
            result.ok = False
            result.erros.append("Arquivo vazio")
            return result

        try:
            raw = open(self.filepath, "rb").read()
            det = chardet.detect(raw)
            encoding = det.get("encoding", "utf-8") or "utf-8"
            result.metadata["encoding_detectado"] = encoding

            content = raw.decode(encoding, errors="replace")
            lines = content.splitlines()
            result.metadata["total_linhas"] = len(lines)

            if not lines:
                result.ok = False
                result.erros.append("Arquivo CSV vazio")
                return result

            first_line = lines[0]
            delimiter = ","
            if ";" in first_line and first_line.count(";") >= first_line.count(","):
                delimiter = ";"
            elif "\t" in first_line and first_line.count("\t") >= first_line.count(","):
                delimiter = "\t"
            result.metadata["delimitador"] = repr(delimiter)

            reader = csv.reader(lines, delimiter=delimiter)
            try:
                headers = next(reader)
                headers = [h.strip() for h in headers]
            except StopIteration:
                headers = []

            row_count = sum(1 for _ in reader)
            result.tabelas.append({
                "nome": os.path.basename(self.filepath),
                "registros": row_count,
                "colunas": len(headers),
                "cabecalhos": headers,
            })
            result.total_registros = row_count

            result.integridade.append({
                "check": "cabecalhos",
                "ok": len(headers) > 0,
                "detalhe": f"{len(headers)} coluna(s) detectada(s)",
            })
            result.integridade.append({
                "check": "linhas",
                "ok": row_count > 0,
                "detalhe": f"{row_count} linha(s) de dados",
            })
            result.integridade.append({
                "check": "encoding",
                "ok": True,
                "detalhe": f"Encoding: {encoding}",
            })
            result.integridade.append({
                "check": "estrutura",
                "ok": True,
                "detalhe": "Arquivo CSV legível",
            })

        except Exception as e:
            result.ok = False
            result.erros.append(f"Erro ao analisar CSV: {str(e)}")

        return result


def analisar_arquivo(filepath, tipo_fonte):
    analyzers = {
        "sqlite": SQLiteAnalyzer,
        "sql": SQLAnalyzer,
        "excel": ExcelAnalyzer,
        "csv": CSVAnalyzer,
    }
    cls = analyzers.get(tipo_fonte)
    if not cls:
        r = AnalysisResult()
        r.ok = False
        r.erros.append(f"Tipo de arquivo não suportado: {tipo_fonte}")
        return r
    return cls(filepath).analyze()


def importar_dados(filepath, tipo_fonte, opcoes=None):
    """Importação real dos dados para o banco ativo."""
    opcoes = opcoes or {}
    db_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "bar_adega.db"
    )
    if not os.path.exists(db_path):
        return {"ok": False, "erro": "Banco de dados alvo não encontrado"}

    if tipo_fonte == "sqlite":
        return _importar_sqlite(filepath, db_path, opcoes)
    elif tipo_fonte == "sql":
        return _importar_sql(filepath, db_path, opcoes)
    elif tipo_fonte == "csv":
        return _importar_csv(filepath, db_path, opcoes)
    elif tipo_fonte == "excel":
        return _importar_excel(filepath, db_path, opcoes)
    else:
        return {"ok": False, "erro": f"Tipo de fonte não suportado: {tipo_fonte}"}


IMPORT_ORDER = [
    "empresa", "usuarios", "categorias", "garcons", "mesas", "produtos",
    "clientes", "caixas", "comandas", "vendas", "itens_comanda",
    "itens_venda", "movimentacoes", "fiado", "contas_pagar",
    "suprimento_sangria", "historico_transferencias", "auditoria",
    "login_attempts",
]


def _get_target_columns(target_conn, table_name):
    try:
        cols = target_conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
        return [c[1] for c in cols]
    except Exception:
        return []


def _importar_sqlite(src_path, dst_path, opcoes):
    t0 = time.time()
    report = {
        "ok": True,
        "tipo_fonte": "sqlite",
        "arquivo": src_path,
        "tabelas_importadas": {},
        "total_registros": 0,
        "erros": [],
        "avisos": [],
        "tempo_importacao_s": 0,
    }
    try:
        src_conn = sqlite3.connect(f"file:{src_path}?mode=ro", uri=True)
        src_conn.row_factory = sqlite3.Row
    except Exception as e:
        return {"ok": False, "erro": f"Erro ao abrir fonte: {str(e)}"}

    dst_conn = sqlite3.connect(dst_path)
    dst_conn.execute("PRAGMA journal_mode=WAL")
    dst_conn.execute("PRAGMA foreign_keys=OFF")

    src_tables = [r[0] for r in src_conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
    ).fetchall()]

    ordered_tables = [t for t in IMPORT_ORDER if t in src_tables]
    remaining = [t for t in src_tables if t not in ordered_tables]
    ordered_tables.extend(remaining)

    try:
        for tname in ordered_tables:
            try:
                src_cols_info = src_conn.execute(f'PRAGMA table_info("{tname}")').fetchall()
                src_cols = [c[1] for c in src_cols_info]
                target_cols = _get_target_columns(dst_conn, tname)
                if not target_cols:
                    report["avisos"].append(f"Tabela '{tname}' não existe no destino — ignorada")
                    continue
                common_cols = [c for c in src_cols if c in target_cols]
                if not common_cols:
                    report["avisos"].append(f"Nenhuma coluna em comum para '{tname}' — ignorada")
                    continue
                rows = src_conn.execute(f'SELECT * FROM "{tname}"').fetchall()
                if not rows:
                    report["tabelas_importadas"][tname] = {"registros": 0, "colunas": common_cols}
                    continue
                placeholders = ",".join(["?"] * len(common_cols))
                col_str = ",".join(f'"{c}"' for c in common_cols)
                inserted = 0
                for row in rows:
                    values = []
                    for c in common_cols:
                        idx = src_cols.index(c)
                        values.append(row[idx])
                    try:
                        dst_conn.execute(
                            f'INSERT OR IGNORE INTO "{tname}" ({col_str}) VALUES ({placeholders})',
                            values
                        )
                        inserted += 1
                    except Exception as e:
                        report["erros"].append(f"{tname}: {str(e)[:100]}")
                report["tabelas_importadas"][tname] = {
                    "registros": inserted,
                    "colunas": common_cols,
                }
                report["total_registros"] += inserted
            except Exception as e:
                report["erros"].append(f"Erro na tabela '{tname}': {str(e)[:100]}")
        dst_conn.execute("PRAGMA foreign_keys=ON")
        dst_conn.commit()
    except Exception as e:
        dst_conn.rollback()
        report["ok"] = False
        report["erros"].append(f"Erro fatal: {str(e)}")
    finally:
        src_conn.close()
        dst_conn.close()
    report["tempo_importacao_s"] = round(time.time() - t0, 2)
    return report


def _importar_sql(src_path, dst_path, opcoes):
    t0 = time.time()
    report = {
        "ok": True,
        "tipo_fonte": "sql",
        "arquivo": src_path,
        "tabelas_importadas": {},
        "total_registros": 0,
        "erros": [],
        "avisos": [],
        "tempo_importacao_s": 0,
    }
    try:
        import chardet
        with open(src_path, "rb") as f:
            raw = f.read()
        enc = chardet.detect(raw).get("encoding", "utf-8") or "utf-8"
        content = raw.decode(enc, errors="replace")
    except Exception as e:
        return {"ok": False, "erro": f"Erro ao ler SQL: {str(e)}"}

    dst_conn = sqlite3.connect(dst_path)
    dst_conn.execute("PRAGMA journal_mode=WAL")
    dst_conn.execute("PRAGMA foreign_keys=OFF")
    insert_pattern = re.compile(
        r"INSERT\s+INTO\s+[`\"']?(\w+)[`\"']?\s*(?:\(([^)]+)\))?\s*VALUES\s*\((.+?)\)\s*;",
        re.IGNORECASE | re.DOTALL,
    )
    table_counts = {}
    try:
        for match in insert_pattern.finditer(content):
            tname = match.group(1)
            col_str = match.group(2)
            vals_str = match.group(3)
            cols = [c.strip().strip('`"\'[]') for c in col_str.split(",")] if col_str else []
            try:
                dst_conn.execute(
                    f'INSERT OR IGNORE INTO "{tname}" ({",".join(f"{c}" for c in cols)}) VALUES ({vals_str})'
                    if cols else
                    f'INSERT OR IGNORE INTO "{tname}" VALUES ({vals_str})'
                )
                table_counts[tname] = table_counts.get(tname, 0) + 1
                report["total_registros"] += 1
            except Exception as e:
                report["erros"].append(f"{tname}: {str(e)[:100]}")
        for tname, count in table_counts.items():
            report["tabelas_importadas"][tname] = {"registros": count}
        dst_conn.execute("PRAGMA foreign_keys=ON")
        dst_conn.commit()
    except Exception as e:
        dst_conn.rollback()
        report["ok"] = False
        report["erros"].append(f"Erro fatal: {str(e)}")
    finally:
        dst_conn.close()
    report["tempo_importacao_s"] = round(time.time() - t0, 2)
    return report


def _importar_csv(src_path, dst_path, opcoes):
    t0 = time.time()
    target_table = opcoes.get("tabela") or opcoes.get("table")
    report = {
        "ok": True,
        "tipo_fonte": "csv",
        "arquivo": src_path,
        "tabelas_importadas": {},
        "total_registros": 0,
        "erros": [],
        "avisos": [],
        "tempo_importacao_s": 0,
    }
    try:
        import chardet
        with open(src_path, "rb") as f:
            raw = f.read()
        enc = chardet.detect(raw).get("encoding", "utf-8") or "utf-8"
        content = raw.decode(enc, errors="replace")
        lines = content.splitlines()
        if not lines:
            return {"ok": False, "erro": "Arquivo CSV vazio"}
        delimiter = ","
        if ";" in lines[0] and lines[0].count(";") >= lines[0].count(","):
            delimiter = ";"
        reader = csv.reader(lines, delimiter=delimiter)
        headers = [h.strip().strip('`"\'[]') for h in next(reader)]
        rows = list(reader)
    except Exception as e:
        return {"ok": False, "erro": f"Erro ao ler CSV: {str(e)}"}

    if not target_table:
        for tname in IMPORT_ORDER:
            if tname in [h.lower() for h in headers] or tname.replace("_", "") in "".join(h.lower() for h in headers):
                target_table = tname
                break
        if not target_table:
            target_table = "clientes"

    dst_conn = sqlite3.connect(dst_path)
    dst_conn.execute("PRAGMA journal_mode=WAL")
    dst_conn.execute("PRAGMA foreign_keys=OFF")
    target_cols = _get_target_columns(dst_conn, target_table)
    if not target_cols:
        dst_conn.close()
        return {"ok": False, "erro": f"Tabela '{target_table}' não existe no destino"}
    common = [h for h in headers if h.lower() in [c.lower() for c in target_cols]]
    if not common:
        dst_conn.close()
        return {"ok": False, "erro": f"Nenhuma coluna do CSV mapeia para '{target_table}'"}

    col_map = [(csv_col, next(tc for tc in target_cols if tc.lower() == csv_col.lower())) for csv_col in common]
    inserted = 0
    try:
        for row in rows:
            if len(row) < len(headers):
                row.extend([""] * (len(headers) - len(row)))
            values = []
            for csv_col, target_col in col_map:
                idx = headers.index(csv_col)
                val = row[idx] if idx < len(row) else ""
                values.append(val if val != "" else None)
            placeholders = ",".join(["?"] * len(col_map))
            col_str = ",".join(f'"{tc}"' for _, tc in col_map)
            try:
                dst_conn.execute(
                    f'INSERT OR IGNORE INTO "{target_table}" ({col_str}) VALUES ({placeholders})',
                    values
                )
                inserted += 1
            except Exception as e:
                report["erros"].append(f"Linha {inserted+1}: {str(e)[:80]}")
        dst_conn.execute("PRAGMA foreign_keys=ON")
        dst_conn.commit()
    except Exception as e:
        dst_conn.rollback()
        report["ok"] = False
        report["erros"].append(f"Erro fatal: {str(e)}")
    finally:
        dst_conn.close()
    report["tabelas_importadas"][target_table] = {
        "registros": inserted,
        "colunas": [tc for _, tc in col_map],
    }
    report["total_registros"] = inserted
    report["tempo_importacao_s"] = round(time.time() - t0, 2)
    return report


def _importar_excel(src_path, dst_path, opcoes):
    t0 = time.time()
    report = {
        "ok": True,
        "tipo_fonte": "excel",
        "arquivo": src_path,
        "tabelas_importadas": {},
        "total_registros": 0,
        "erros": [],
        "avisos": [],
        "tempo_importacao_s": 0,
    }
    try:
        from openpyxl import load_workbook
        wb = load_workbook(src_path, read_only=True, data_only=True)
    except ImportError:
        return {"ok": False, "erro": "Biblioteca openpyxl não instalada"}
    except Exception as e:
        return {"ok": False, "erro": f"Erro ao abrir Excel: {str(e)}"}

    dst_conn = sqlite3.connect(dst_path)
    dst_conn.execute("PRAGMA journal_mode=WAL")
    dst_conn.execute("PRAGMA foreign_keys=OFF")

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows_iter = ws.iter_rows(values_only=True)
        all_rows = list(rows_iter)
        if not all_rows:
            continue
        headers = [str(c).strip().strip('`"\'[]') if c is not None else "" for c in all_rows[0]]
        data_rows = all_rows[1:]
        target_table = None
        for tname in IMPORT_ORDER:
            if tname == sname.lower().strip() or tname.replace("_", "") == sname.lower().replace(" ", "").replace("_", ""):
                target_table = tname
                break
        if not target_table:
            target_table = sname.lower().strip().replace(" ", "_")
        target_cols = _get_target_columns(dst_conn, target_table)
        if not target_cols:
            report["avisos"].append(f"Aba '{sname}' → tabela '{target_table}' não existe — ignorada")
            continue
        common = [h for h in headers if h.lower() in [c.lower() for c in target_cols]]
        if not common:
            report["avisos"].append(f"Aba '{sname}' — nenhuma coluna mapeável — ignorada")
            continue
        col_map = [(csv_col, next(tc for tc in target_cols if tc.lower() == csv_col.lower())) for csv_col in common]
        inserted = 0
        for row in data_rows:
            values = []
            for csv_col, target_col in col_map:
                idx = headers.index(csv_col)
                val = row[idx] if idx < len(row) else None
                if val == "":
                    val = None
                values.append(val)
            placeholders = ",".join(["?"] * len(col_map))
            col_str = ",".join(f'"{tc}"' for _, tc in col_map)
            try:
                dst_conn.execute(
                    f'INSERT OR IGNORE INTO "{target_table}" ({col_str}) VALUES ({placeholders})',
                    values
                )
                inserted += 1
            except Exception as e:
                report["erros"].append(f"Aba '{sname}' linha {inserted+1}: {str(e)[:80]}")
        report["tabelas_importadas"][target_table] = {
            "registros": inserted,
            "colunas": [tc for _, tc in col_map],
        }
        report["total_registros"] += inserted

    try:
        dst_conn.execute("PRAGMA foreign_keys=ON")
        dst_conn.commit()
    except Exception as e:
        dst_conn.rollback()
        report["ok"] = False
        report["erros"].append(f"Erro ao commitar: {str(e)}")
    finally:
        dst_conn.close()
        wb.close()
    report["tempo_importacao_s"] = round(time.time() - t0, 2)
    return report
